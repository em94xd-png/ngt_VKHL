from openpyxl import *
import os, xml.etree.ElementTree, sys, pyautogui
from datetime import *
from openpyxl.styles import PatternFill
import time

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import script_config

current_path = os.path.dirname(os.path.abspath(__file__))

excel_file = "Briefing.xlsx"
path_excel = os.path.join(current_path, excel_file)

arrival = "res_detail_142956722.XML"
path_arrival = os.path.join(current_path, arrival)

tree = xml.etree.ElementTree.parse(path_arrival)
root = tree.getroot()

arr_row = 1

for _ in root.findall(".//G_RESERVATION"):
    arr_row += 1

pyautogui.FAILSAFE = True

os.startfile(path_excel)
time.sleep(1.5)
pyautogui.hotkey("ctrl", "g")
pyautogui.write("a16")
pyautogui.press("enter")
pyautogui.hotkey("shift", "space")
time.sleep(.15)
for _ in range(arr_row - 2):
    pyautogui.PAUSE = 0.001
    pyautogui.hotkey("ctrl", "c")
    pyautogui.hotkey("ctrl", "+")

departure = "departure_all_142956320.XML"
path_departure = os.path.join(current_path, departure)

tree = xml.etree.ElementTree.parse(path_departure)
root = tree.getroot()

dep_row = arr_row

for _ in root.findall(".//G_ROOM"):
    dep_row += 1

pyautogui.hotkey("ctrl", "g")
pyautogui.write(f"a{arr_row + 17}")
pyautogui.press("enter")
pyautogui.hotkey("shift", "space")
time.sleep(.15)
for _ in range((dep_row - arr_row) - 1):
    pyautogui.PAUSE = 0.001
    pyautogui.hotkey("ctrl", "c")
    pyautogui.hotkey("ctrl", "+")
pyautogui.hotkey("ctrl", "home")
pyautogui.press("down", presses=2)
pyautogui.hotkey("ctrl", "s")
pyautogui.hotkey("alt", "f4")

while True:
    if pyautogui.pixelMatchesColor(710, 523, (12, 12, 12)):
        break

wb = load_workbook(path_excel)

ws = "Ori Format"
ws = wb[ws]

ws["A2"] = f"Daily Briefing\n{script_config.td_full_d} {script_config.td_full_m} {script_config.td_short_date}, {script_config.td_yyyy}"

history_forecast = "history_forecast_142950236.XML"
path_history_forecast = os.path.join(current_path, history_forecast)

tree = xml.etree.ElementTree.parse(path_history_forecast)
root = tree.getroot()

for _ in root.findall(".//G_GPAGEID/LIST_G_REC_TYPE/G_REC_TYPE/LIST_G_CONSIDERED_DATE/G_CONSIDERED_DATE"):
    hu = _.find("HOUSE_USE_ROOMS").text if _.find("HOUSE_USE_ROOMS").text is not None else "N/A"
    oc = round(float(_.find("CF_OCCUPANCY").text), 2) if _.find("CF_OCCUPANCY").text is not None else "N/A"
    arr = _.find("ARRIVAL_ROOMS").text if _.find("ARRIVAL_ROOMS").text is not None else "-"
    dep = _.find("DEPARTURE_ROOMS").text if _.find("DEPARTURE_ROOMS").text is not None else "-"
    ppl = _.find("NO_PERSONS").text if _.find("NO_PERSONS").text is not None else "-"
    if _.find("CONSIDERED_DATE").text == "25-JUL-26":
        ws["E5"] = arr
        ws["E7"] = dep
    if _.find("CONSIDERED_DATE").text == "26-JUL-26":
        ws["M5"] = arr
        ws["M7"] = dep
    if _.find("CONSIDERED_DATE").text == "27-JUL-26":
        ws["O5"] = arr
        ws["O7"] = dep
    if _.find("CONSIDERED_DATE").text == "28-JUL-26":
        ws["Q5"] = arr
        ws["Q7"] = dep
    if _.find("CONSIDERED_DATE").text == "29-JUL-26":
        ws["S5"] = arr
        ws["S7"] = dep
    if _.find("CONSIDERED_DATE").text == "30-JUL-26":
        ws["U5"] = arr
        ws["U7"] = dep

history_forecast_AVC = "history_forecast_142950252.XML"
path_history_forecast_AVC = os.path.join(current_path, history_forecast_AVC)

tree = xml.etree.ElementTree.parse(path_history_forecast_AVC)
root = tree.getroot()

for _ in root.findall(".//G_GPAGEID/LIST_G_REC_TYPE/G_REC_TYPE/LIST_G_CONSIDERED_DATE/G_CONSIDERED_DATE"):
    cmp = _.find("COMPLIMENTARY_ROOMS").text if _.find("COMPLIMENTARY_ROOMS").text is not None else "-"
    if cmp == "0":
        cmp = "-"
    hu = int(_.find("HOUSE_USE_ROOMS").text if str(_.find("HOUSE_USE_ROOMS").text) else 0)
    if hu == "0":
        hu = "-"
    ns = _.find("NO_SHOW_ROOMS").text if _.find("NO_SHOW_ROOMS").text else "-"
    if ns == "0":
        ns = "-"
    adr = round(float(_.find("CF_AVERAGE_ROOM_RATE").text), 2) if _.find("CF_AVERAGE_ROOM_RATE").text is not None else "-"
    arr = _.find("ARRIVAL_ROOMS").text if _.find("ARRIVAL_ROOMS").text is not None else "-"
    dep = _.find("DEPARTURE_ROOMS").text if _.find("DEPARTURE_ROOMS").text is not None else "-"
    rm = int(_.find("NO_ROOMS").text if str(_.find("NO_ROOMS").text) else 0)
    ppl = _.find("NO_PERSONS").text if _.find("NO_PERSONS").text is not None else "-"
    if _.find("CONSIDERED_DATE").text == "24-JUL-26":
        ws["B5"] = cmp
        ws["B7"] = hu
        ws["B9"] = ns
        ws["B11"] = adr
        ws["B11"].number_format = "#,###.##"
        ws["B12"] = oc
        ws["B12"].number_format = "#.##\"%\""
    if _.find("CONSIDERED_DATE").text == "25-JUL-26":
        ws["D4"] = "25-Jul"
        ws["F5"] = arr
        ws["F7"] = dep
        ws["E9"] = rm
        ws["G12"] = ppl
    if _.find("CONSIDERED_DATE").text == "26-JUL-26":
        ws["N5"] = arr
        ws["N7"] = dep
        ws["M9"] = rm
        ws["M11"] = (rm - hu) / (327 - hu) * 100
        ws["M12"] = ppl
    if _.find("CONSIDERED_DATE").text == "27-JUL-26":
        ws["P5"] = arr
        ws["P7"] = dep
        ws["O9"] = rm
        ws["O11"] = (rm - hu) / (327 - hu) * 100
        ws["O12"] = ppl
    if _.find("CONSIDERED_DATE").text == "28-JUL-26":
        ws["R5"] = arr
        ws["R7"] = dep
        ws["Q9"] = rm
        ws["Q11"] = (rm - hu) / (327 - hu) * 100
        ws["Q12"] = ppl
    if _.find("CONSIDERED_DATE").text == "29-JUL-26":
        ws["T5"] = arr
        ws["T7"] = dep
        ws["S9"] = rm
        ws["S11"] = (rm - hu) / (327 - hu) * 100
        ws["S12"] = ppl
    if _.find("CONSIDERED_DATE").text == "30-JUL-26":
        ws["V5"] = arr
        ws["V7"] = dep
        ws["U9"] = rm
        ws["U11"] = (rm - hu) / (327 - hu) * 100
        ws["U12"] = ppl

cmp = int(ws["B5"].value if str(ws["B5"].value) else 0)
hu = int(ws["B7"].value if str(ws["B7"].value) else 0)
rm = int(ws["E9"].value if str(ws["E9"].value) else 0)

ws["E11"] = round(float((rm - (cmp + hu)) / 327 * 100), 2)
ws["E11"].number_format = "#.##\"%\""

split_adult_child = "resfutureoccupancy_142949802.XML"
path_split_adult_child = os.path.join(current_path, split_adult_child)

tree = xml.etree.ElementTree.parse(path_split_adult_child)
root = tree.getroot()

for _ in root.findall(".//G_RESV_TYPE"):
    adl = _.find("SUMADULTS").text if _.find("SUMADULTS").text is not None else "-"
    chd = _.find("SUMCHILDREN").text if _.find("SUMCHILDREN").text is not None else "-"
    if _.find(".//D_DATE").text == "25-JUL-26":
        ws["E12"] = adl
        ws["E13"] = chd

room_upgrade = "finjrnlbytrans_142949822.XML"
path_room_upgrade = os.path.join(current_path, room_upgrade)

tree = xml.etree.ElementTree.parse(path_room_upgrade)
root = tree.getroot()

for _ in root:
    if _.tag == "R_DEBIT":
        val = float(_.text)
        if _.text == "0":
            ws["I5"] = "-"
        else:
            if val.is_integer():
                ws["I5"] = int(val)
                ws["I5"].number_format = "#,##0"
            else:
                ws["I5"] = round(val, 2)
                ws["I5"].number_format = "#,##0.00"

room_upgrade_mtd = "finjrnlbytrans_142950555.XML"
path_room_upgrade_mtd = os.path.join(current_path, room_upgrade_mtd)

tree = xml.etree.ElementTree.parse(path_room_upgrade_mtd)
root = tree.getroot()

for _ in root:
    if _.tag == "R_DEBIT":
        val = float(_.text)
        if _.text == "0":
            ws["J5"] = "-"
        else:
            if val.is_integer():
                ws["J5"] = int(val)
                ws["J5"].number_format = "#,##0"
            else:
                ws["J5"] = round(val, 2)
                ws["J5"].number_format = "#,##0.00"

late_checkout = "finjrnlbytrans_142950099.XML"
path_late_checkout = os.path.join(current_path, late_checkout)

tree = xml.etree.ElementTree.parse(path_late_checkout)
root = tree.getroot()

for _ in root:
    if _.tag == "R_DEBIT":
        val = float(_.text)
        if _.text == "0":
            ws["I7"] = "-"
        else:
            if val.is_integer():
                ws["I7"] = int(val)
                ws["I7"].number_format = "#,##0"
            else:
                ws["I7"] = round(val, 2)
                ws["I7"].number_format = "#,##0.00"

late_checkout_mtd = "finjrnlbytrans_142950190.XML"
path_late_checkout_mtd = os.path.join(current_path, late_checkout_mtd)

tree = xml.etree.ElementTree.parse(path_late_checkout_mtd)
root = tree.getroot()

for _ in root:
    if _.tag == "R_DEBIT":
        val = float(_.text)
        if _.text == "0":
            ws["J7"] = "-"
        else:
            if val.is_integer():
                ws["J7"] = int(val)
                ws["J7"].number_format = "#,##0"
            else:
                ws["J7"] = round(val, 2)
                ws["J7"].number_format = "#,##0.00"

tour_commission = "finjrnlbytrans_142950122.XML"
path_tour_commission = os.path.join(current_path, tour_commission)

tree = xml.etree.ElementTree.parse(path_tour_commission)
root = tree.getroot()

for _ in root:
    if _.tag == "R_DEBIT":
        val = float(_.text)
        if _.text == "0":
            ws["I9"] = "-"
        else:
            if val.is_integer():
                ws["I9"] = int(val)
                ws["I9"].number_format = "#,##0"
            else:
                ws["I9"] = round(val, 2)
                ws["I9"].number_format = "#,##0.00"

tour_commission_mtd = "finjrnlbytrans_142950546.XML"
path_tour_commission_mtd = os.path.join(current_path, tour_commission_mtd)

tree = xml.etree.ElementTree.parse(path_tour_commission_mtd)
root = tree.getroot()

for _ in root:
    if _.tag == "R_DEBIT":
        val = float(_.text)
        if _.text == "0":
            ws["J9"] = "-"
        else:
            if val.is_integer():
                ws["J9"] = int(val)
                ws["J9"].number_format = "#,##0"
            else:
                ws["J9"] = round(val, 2)
                ws["J9"].number_format = "#,##0.00"

gift_shop = "finjrnlbytrans_142950142.XML"
path_gift_shop = os.path.join(current_path, gift_shop)

tree = xml.etree.ElementTree.parse(path_gift_shop)
root = tree.getroot()

for _ in root:
    if _.tag == "R_DEBIT":
        val = float(_.text)
        if _.text == "0":
            ws["I12"] = "-"
        else:
            if val.is_integer():
                ws["I12"] = int(val)
                ws["I12"].number_format = "#,##0"
            else:
                ws["I12"] = round(val, 2)
                ws["I12"].number_format = "#,##0.00"

gift_shop_mtd = "finjrnlbytrans_142949873.XML"
path_gift_shop_mtd = os.path.join(current_path, gift_shop_mtd)

tree = xml.etree.ElementTree.parse(path_gift_shop_mtd)
root = tree.getroot()

for _ in root:
    if _.tag == "R_DEBIT":
        val = float(_.text)
        if _.text == "0":
            ws["J12"] = "-"
        else:
            if val.is_integer():
                ws["J12"] = int(val)
                ws["J12"].number_format = "#,##0"
            else:
                ws["J12"] = round(val, 2)
                ws["J12"].number_format = "#,##0.00"

pf = ("PAID", "CTC", "POA", "COA", "POD", "COMP", "MASTER", "COD")

red_color = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

arrival = "res_detail_142956722.XML"
path_arrival = os.path.join(current_path, arrival)

tree = xml.etree.ElementTree.parse(path_arrival)
root = tree.getroot()

arr_start = 16

for _ in root.findall(".//G_RESERVATION"):
    name = _.find("FULL_NAME_NO_SHR_IND").text if _.find("FULL_NAME_NO_SHR_IND") is not None else "-"
    rm = _.find("DISP_ROOM_NO").text if _.find("DISP_ROOM_NO") is not None else "-"
    ta = _.find("COMPANY_NAME").text if _.find("COMPANY_NAME") is not None else "-"
    vip = _.find("VIP").text if _.find("VIP") is not None else "-"
    memb = None
    for mb in _.findall(".//G_MEM_TYPE_LEVEL"):
        memb = mb.find("MEMBERSHIP_LEVEL").text
    adl = _.find("ADULTS").text if _.find("ADULTS") is not None else "-"
    chd = _.find("CHILDREN").text if _.find("CHILDREN") is not None else "-"
    arr = _.find("ARRIVAL").text if _.find("ARRIVAL") is not None else "-"
    dep = _.find("DEPARTURE").text if _.find("DEPARTURE") is not None else "-"
    eta = _.find("ARRIVAL_TIME").text if _.find("ARRIVAL_TIME") is not None else "-"
    cmt = None
    for cmt_res in _.findall(".//G_COMMENT_RESV_NAME_ID"):
        in_cmt = cmt_res.find("RES_COMMENT").text
        if in_cmt and in_cmt.startswith(pf):
            cmt = in_cmt
        if cmt is None:
            if cmt_res.find("RES_COMMENT_DESCRIPTION").text == "RESERVATION":
                for _ in str(pf).strip().lower():
                        if _ in str(in_cmt).strip().lower():
                            cmt = in_cmt

    set_name = [ _.strip() for _ in name.split(",")]
    ln, fn, tt = set_name[0], set_name[1], set_name[2]
    if (ln, fn, tt):
        ws[f"A{arr_start}"] = f"{tt}. {fn} {ln}"
        ws[f"A{arr_start}"].fill = PatternFill(fill_type=None)
    if tt is None:
        ws[f"A{arr_start}"] = f"{tt}. {fn} {ln}"
        ws[f"A{arr_start}"].fill = red_color

    if "Maintenance".strip().lower() in str(name).strip().lower():
        for _ in ws[f"A{arr_start}:P{arr_start}"][0]:
            _.fill = red_color

    ws[f"C{arr_start}"] = "-"

    ws[f"D{arr_start}"] = rm

    if "MI Squared".strip().lower() in str(cmt).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nMI Squared"
    if "Booking.com".strip().lower() in str(ta).strip().lower():
            ws[f"D{arr_start}"] = f"{rm}\nBooking.com"
    if "BG Asia".strip().lower() in str(ta).strip().lower():
            ws[f"D{arr_start}"] = f"{rm}\nBG Asia"
    if "Siam Tours".strip().lower() in str(ta).strip().lower():
            ws[f"D{arr_start}"] = f"{rm}\nSiam Tours"
    if "Travelbullz".strip().lower() in str(ta).strip().lower():
            ws[f"D{arr_start}"] = f"{rm}\nTravelbullz"
    if "BTC".strip().lower() in str(ta).strip().lower():
            ws[f"D{arr_start}"] = f"{rm}\nBTC"
    if "DERTOUR".strip().lower() in str(ta).strip().lower():
            ws[f"D{arr_start}"] = f"{rm}\nGo Vacation"
    if "Expedia".strip().lower() in str(ta).strip().lower():
            ws[f"D{arr_start}"] = f"{rm}\nExpedia"
    if "DNATA".strip().lower() in str(ta).strip().lower():
            ws[f"D{arr_start}"] = f"{rm}\nDNATA"
    if "Miki".strip().lower() in str(ta).strip().lower():
            ws[f"D{arr_start}"] = f"{rm}\nMiki Travel"
    if "Rak".strip().lower() in str(ta).strip().lower():
            ws[f"D{arr_start}"] = f"{rm}\nRak Journeys"
    if "Hotelbeds".strip().lower() in str(ta).strip().lower():
            ws[f"D{arr_start}"] = f"{rm}\nHotelbeds"

    if vip is not None:
        ws[f"F{arr_start}"] = vip
        if memb:
            ws[f"F{arr_start}"] = f"{vip}\n{memb}"
    if vip is None:
        ws[f"F{arr_start}"] = "-"
        if memb in ["SILVER", "GOLD"]:
            if diff_date.days < 7:
                if vip not in ["VIP1", "VIP2", "VIPS"]:
                    ws[f"F{arr_start}"] = f"VIPG\n{memb}"
        if memb in ["PLATINUM"]:
            if vip not in ["VIP1", "VIP2"]:
                ws[f"F{arr_start}"] = f"VIPS\n{memb}"
        if memb in ["TITANIUM", "RED"]:
            if vip not in ["VIP1"]:
                ws[f"F{arr_start}"] = f"VIP2\n{memb}"

    arr_date = datetime.strptime(arr, "%d/%m/%y")
    dep_date = datetime.strptime(dep, "%d/%m/%y")
    diff_date = dep_date - arr_date

    if diff_date.days >= 7:
        if vip not in ["VIP1", "VIP2"]:
            if memb:
                ws[f"F{arr_start}"] = f"VIPS\n{memb}"
            if not memb:
                ws[f"F{arr_start}"] = f"VIPS"

    if memb in ["SILVER", "GOLD"]:
        if diff_date.days < 7:
            if vip not in ["VIP1", "VIP2", "VIPS"]:
                ws[f"F{arr_start}"] = f"VIPG\n{memb}"
    if memb in ["PLATINUM"]:
        if vip not in ["VIP1", "VIP2"]:
            ws[f"F{arr_start}"] = f"VIPS\n{memb}"
    if memb in ["TITANIUM", "RED"]:
        if vip not in ["VIP1"]:
            ws[f"F{arr_start}"] = f"VIP2\n{memb}"

    if not "MI Squared".strip().lower() in str(cmt).strip().lower():
        if "JADE/RUBY".strip().lower() in str(cmt).strip().lower():
            ws[f"F{arr_start}"] = f"VIPO\nJADE"
        if "DIAMOND".strip().lower() in str(cmt).strip().lower():
            ws[f"F{arr_start}"] = f"VIPO\nDIAMOND"
        if "PLATINUM".strip().lower() in str(cmt).strip().lower():
            ws[f"F{arr_start}"] = f"VIPO\nPLATINUM"
        if "ROYAL".strip().lower() in str(cmt).strip().lower():
            ws[f"F{arr_start}"] = f"VIPO\nROYAL"
    if "MI Squared".strip().lower() in str(cmt).strip().lower():
        if diff_date.days < 7:
            ws[f"F{arr_start}"] = "-"
        if diff_date.days >= 7:
            ws[f"F{arr_start}"] = "VIPS"

    if chd is not None:
        ws[f"H{arr_start}"] = f"{adl}+{chd}"
    if chd is None or chd == "0":
        ws[f"H{arr_start}"] = f"{adl}"

    ws[f"I{arr_start}"] = arr

    ws[f"J{arr_start}"] = dep

    ws[f"K{arr_start}"] = "-"

    if eta is not None:
        ws[f"L{arr_start}"] = eta
        ws[f"L{arr_start}"].value = eta.replace(":", ".")
    if eta is None:
        ws[f"L{arr_start}"] = "15.00"

    ws[f"N{arr_start}"] = "OWN"

    ws[f"P{arr_start}"] = cmt

    arr_start += 1

departure = "departure_all_142956320.XML"
path_departure = os.path.join(current_path, departure)

tree = xml.etree.ElementTree.parse(path_departure)
root = tree.getroot()

dep_start = arr_start + 2

for _ in root.findall(".//G_ROOM"):
    name = _.find("GUEST_NAME").text if _.find("GUEST_NAME") is not None else "-"
    rm = _.find("ROOM").text if _.find("ROOM") is not None else "-"
    ta = _.find("TRAVEL_AGENT_NAME").text if _.find("TRAVEL_AGENT_NAME") is not None else "-"
    vip = _.find("VIP").text if _.find("VIP") is not None else "-"
    memb = None
    for mb in _.findall(".//LIST_G_MEMBERSHIP/G_MEMBERSHIP"):
        memb = mb.find("MEMBERSHIP_LEVEL").text
    adl = _.find("ADULTS").text if _.find("ADULTS") is not None else "-"
    chd = _.find("CHILDREN").text if _.find("CHILDREN") is not None else "-"
    arr = _.find("CHAR_ARRIVAL").text if _.find("CHAR_ARRIVAL") is not None else "-"
    dep = _.find("CHAR_DEPART").text if _.find("CHAR_DEPART") is not None else "-"
    etd = _.find("DEPARTURE_TIME").text if _.find("DEPARTURE_TIME") is not None else "-"
    cmt = None
    for cmt_res in _.findall("LIST_G_COMMENT_RESV_NAME_ID/G_COMMENT_RESV_NAME_ID"):
        in_cmt = cmt_res.find("RES_COMMENT").text
        if in_cmt and in_cmt.startswith(pf):
            cmt = in_cmt
        if cmt is None:
            if cmt_res.find("RES_COMMENT_DESCRIPTION").text == "Reservation":
                for _ in str(pf).strip().lower():
                    if _ in str(in_cmt).strip().lower():
                        cmt = in_cmt

    set_name = [ _.strip() for _ in name.split(",")]
    ln, fn, tt = set_name[0], set_name[1], set_name[2]
    if (ln, fn, tt):
        ws[f"A{dep_start}"] = f"{tt}. {fn} {ln}"
        ws[f"A{dep_start}"].fill = PatternFill(fill_type=None)
    if tt is None:
        ws[f"A{dep_start}"] = f"{tt}. {fn} {ln}"
        ws[f"A{dep_start}"].fill = red_color

    if "Maintenance".strip().lower() in str(name).strip().lower():
        for _ in ws[f"A{dep_start}:P{arr_start}"][0]:
            _.fill = red_color

    ws[f"C{dep_start}"] = "-"

    ws[f"D{dep_start}"] = rm

    if "MI Squared".strip().lower() in str(cmt).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nMI Squared"
    if "Booking.com".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nBooking.com"
    if "BG Asia".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nBG Asia"
    if "Siam Tours".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nSiam Tours"
    if "Travelbullz".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nTravelbullz"
    if "BTC".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nBTC"
    if "DERTOUR".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nGo Vacation"
    if "Expedia".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nExpedia"
    if "DNATA".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nDNATA"
    if "Miki".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nMiki Travel"
    if "Rak".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nRak Journeys"
    if "Hotelbeds".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nHotelbeds"

    if vip is not None:
        ws[f"F{dep_start}"] = vip
        if memb:
            ws[f"F{dep_start}"] = f"{vip}\n{memb}"
    if vip is None:
        ws[f"F{dep_start}"] = "-"
        if memb in ["SILVER", "GOLD"]:
            if diff_date.days < 7:
                if vip not in ["VIP1", "VIP2", "VIPS"]:
                    ws[f"F{dep_start}"] = f"VIPG\n{memb}"
        if memb in ["PLATINUM"]:
            if vip not in ["VIP1", "VIP2"]:
                ws[f"F{dep_start}"] = f"VIPS\n{memb}"
        if memb in ["TITANIUM", "RED"]:
            if vip not in ["VIP1"]:
                ws[f"F{dep_start}"] = f"VIP2\n{memb}"

    arr_date = datetime.strptime(arr, "%d/%m/%y")
    dep_date = datetime.strptime(dep, "%d/%m/%y")
    diff_date = dep_date - arr_date

    if diff_date.days >= 7:
        if vip not in ["VIP1", "VIP2"]:
            if memb:
                ws[f"F{dep_start}"] = f"VIPS\n{memb}"
            if not memb:
                ws[f"F{dep_start}"] = f"VIPS"

    if memb in ["SILVER", "GOLD"]:
        if diff_date.days < 7:
            if vip not in ["VIP1", "VIP2", "VIPS"]:
                ws[f"F{dep_start}"] = f"VIPG\n{memb}"
    if memb in ["PLATINUM"]:
        if vip not in ["VIP1", "VIP2"]:
            ws[f"F{dep_start}"] = f"VIPS\n{memb}"
    if memb in ["TITANIUM", "RED"]:
        if vip not in ["VIP1"]:
            ws[f"F{dep_start}"] = f"VIP2\n{memb}"

    if not "MI Squared".strip().lower() in str(cmt).strip().lower():
        if "JADE/RUBY".strip().lower() in str(cmt).strip().lower():
            ws[f"F{dep_start}"] = f"VIPO\nJADE"
        if "DIAMOND".strip().lower() in str(cmt).strip().lower():
            ws[f"F{dep_start}"] = f"VIPO\nDIAMOND"
        if "PLATINUM".strip().lower() in str(cmt).strip().lower():
            ws[f"F{dep_start}"] = f"VIPO\nPLATINUM"
        if "ROYAL".strip().lower() in str(cmt).strip().lower():
            ws[f"F{dep_start}"] = f"VIPO\nROYAL"
    if "MI Squared".strip().lower() in str(cmt).strip().lower():
        if diff_date.days < 7:
            ws[f"F{dep_start}"] = "-"
        if diff_date.days >= 7:
            ws[f"F{dep_start}"] = "VIPS"

    if chd is not None:
        ws[f"H{dep_start}"] = f"{adl}+{chd}"
    if chd is None or chd == "0":
        ws[f"H{dep_start}"] = f"{adl}"

    ws[f"I{dep_start}"] = arr

    ws[f"J{dep_start}"] = dep

    ws[f"K{dep_start}"] = "-"

    if etd is not None:
        ws[f"L{dep_start}"] = etd
        ws[f"L{dep_start}"].value = etd.replace(":", ".")
    if etd is None:
        ws[f"L{dep_start}"] = "12.00"

    ws[f"N{dep_start}"] = "OWN"

    ws[f"P{dep_start}"] = cmt

    dep_start += 1

wb.save(path_excel)