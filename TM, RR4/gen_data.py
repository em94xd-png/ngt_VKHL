from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import openpyxl, json, os, shutil, pyautogui, subprocess, pygetwindow, time, pyperclip, sys
import xml.etree.ElementTree
from datetime import datetime, timedelta, date
from urllib.parse import urlparse, parse_qs

site_OPERA = "https://mtca2.oraclehospitality.ap-singapore-1.ocs.oraclecloud.com/MINOR/operacloud/faces/opera-cloud-index/OperaCloud"

def format2_yesterday():
    today = date.today()
    yesterday = today - timedelta(days=1)
    return yesterday.strftime("%d%m")

# Open Opera
subprocess.run(["cmd", "/c", "start", "msedge", site_OPERA])

pygetwindow.getWindowsWithTitle("Opera Cloud")[0].maximize()

# In Opera  
time.sleep(2.5)
pyautogui.hotkey("ctrl", "-", "-", "-", "-", "-", "-", "-", "-", "-", "-",)
pyautogui.hotkey("ctrl", "=", "=", "=",)
time.sleep(5)

# To report search
pyautogui.press("tab", presses=5, interval=0.01)
pyautogui.press("right", presses=6, interval=0.01)
pyautogui.press("down", interval=0.01)
pyautogui.press("enter", interval=0.01)
time.sleep(5)
pyautogui.press("tab", interval=0.01)

# Immigration Report
pyautogui.write("immigration_report", interval=.01)
pyautogui.press("enter", interval=.01)
time.sleep(3)
pyautogui.press("tab", presses=9, interval=.01)
pyautogui.press("down", presses=2, interval=.01)
time.sleep(1.5)
pyautogui.press("right", interval=.01)
pyautogui.press("tab", presses=3, interval=.01)
pyautogui.press("enter", interval=.01)
time.sleep(3.5)
# Immigration Report: Config
pyautogui.hotkey("ctrl", "a", interval=.01)
pyautogui.write(format2_yesterday(), interval=.01)
pyautogui.press("tab", interval=.01)
time.sleep(1)
pyautogui.write("ARRIVAL", interval=.01)
pyautogui.press("tab", presses=2, interval=.01)
pyautogui.press("space", interval=.01)
time.sleep(.5)
pyautogui.press("tab", presses=2, interval=.01)
pyautogui.press("enter", interval=.01)
time.sleep(1)
# Immigration Report: Save
pyautogui.press("tab", presses=2, interval=.01)
pyautogui.press("space", presses=2, interval=.01)
pyautogui.press("tab", presses=2, interval=.01)
time.sleep(.5)
pyautogui.press("enter", interval=.01)
time.sleep(5)
# Immigration Report: Download
pyautogui.hotkey("ctrl", "j", interval=.01)
time.sleep(.5)
pyautogui.hotkey("ctrl", "l", interval=.01)
pyautogui.hotkey("ctrl", "c", interval=.01)

immigration_url = pyperclip.paste()

parse_url = urlparse(immigration_url)
query_url = parse_qs(parse_url.query)

def immigration_id():
     if "rep" in query_url:
          rep_id = query_url["rep"][0]
          split_id = rep_id.split("_")[1]
          return split_id

path_download = os.environ.get("USERPROFILE").__add__(r"\Downloads")
immigration_file = f"immigration_report_{immigration_id()}.XML"

time.sleep(1)
pyautogui.hotkey("ctrl", "j", interval=.01)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)

ytd = date.today() - timedelta(days=1)
ytd_date = ytd.strftime("%d.%m.%y")

data_path = r"\\LMPC202507256L\Keeper\OTH"
data_excel = f"get_{ytd_date}.xlsx"
path_data_excel = os.path.join(data_path, data_excel)

xml_file = os.path.join(path_download, immigration_file)

time.sleep(2.5)

if not os.path.exists(xml_file):
     sys.exit()

tree = xml.etree.ElementTree.parse(xml_file)
root = tree.getroot()

ws1 = "Sheet1"
ws2 = "Sheet2"

wb = load_workbook(os.path.join(path_data_excel))

ws1 = wb[ws1]
ws2 = wb[ws2]

current_path = os.path.dirname(os.path.abspath(__file__))

with open((os.path.join(current_path, "tm_nt.json")), "r", encoding="utf-8") as file:
     tm_nt_json = json.load(file)
with open((os.path.join(current_path, "rr_nt.json")), "r", encoding="utf-8") as file:
     rr_nt_json = json.load(file)
with open((os.path.join(current_path, "rr_ct.json")), "r", encoding="utf-8") as file:
     rr_ct_json = json.load(file)

def sw_date_format(ad, be, sc):
     if ad and ad.strip() != "":
          try:
               ad = datetime.strptime(ad.strip(), "%m/%d/%Y")
               return ad.strftime("%d/%m/%Y")
          except ValueError:
               return ad
     if be and be.strip() != "":
          try:
               be = datetime.strptime(be.strip(), "%m/%d/%Y")
               cvt_be = be.year + 543
               return be.strftime(f"%d/%m/{cvt_be}")
          except ValueError:
               return be
     if sc and str(sc).strip() != "":
          try:
               sc = datetime.strptime(sc, "%d-%m-%Y")
               return sc.strftime("%d/%m/%Y")
          except ValueError:
               return sc

for _ in root.findall(".//G_IMMIGRATION"):
        fn = _.find("FIRST_NAME").text
        ln = _.find("LAST_NAME").text
        gd = _.find("SEX").text
        pn = _.find("PASSPORT").text
        nt = _.find("NATIONALITY").text
        bd = _.find("DATE_OF_BIRTH").text
        bd_ad = sw_date_format(bd, None, None)
        dep = _.find("DEPARTURE_DATE").text
        dep_ad = sw_date_format(dep, None, None)
        dep_be = sw_date_format(None, dep, None)
        rn = _.find("ROOM").text
        tm_data = [fn, None, ln, gd, pn, nt, bd_ad, dep_ad]
        rr_data = [None, None, None, rn, None, None, None, None, gd, fn, None, ln, nt, None, pn, None, None, None, None, nt, nt, None, None, dep_be, None, None]
        ws1.append(tm_data)
        ws2.append(rr_data)

ws3 = wb["Sheet3"]
# wb.active = ws3

ws3_fn_ln_pn = {}
ws3_fn_ln_ct = {}
ws3_fn_ln_bd = {}
for _ in ws3.iter_rows(values_only=True, min_row=2):
    ws3_fn = _[1]
    ws3_ln = _[0]    
    ws3_pn = _[4]     
    ws3_ct = _[3]
    ws3_bd = _[2]
    if ws3_fn and ws3_ln and ws3_pn != "":
          ws3_key = f"{str(ws3_fn).strip().lower()}_{str(ws3_ln).strip().lower()}"
          ws3_fn_ln_pn[ws3_key] = str(ws3_pn).strip()
          ws3_fn_ln_ct[ws3_key] = str(ws3_ct).strip()
          ws3_fn_ln_bd[ws3_key] = str(ws3_bd).strip()

for _ in range(ws1.max_row, 1, -1):
     ws1_fn = ws1.cell(row=_, column=1).value
     ws1_ln = ws1.cell(row=_, column=3).value
     ws1_pn = ws1.cell(row=_, column=5).value
     ws1_nt = ws1.cell(row=_, column=6).value
     ws1_bd = ws1.cell(row=_, column=7).value
     if ws1_fn and ws1_ln != "":
          ws1_key = f"{str(ws1_fn).strip().lower()}_{str(ws1_ln).strip().lower()}"
          if ws1_key in ws3_fn_ln_pn:
               for_empty_passport = ws3_fn_ln_pn[ws1_key]
               if ws1_pn is None or str(ws1_pn).strip() == "":
                    ws1.cell(row=_, column=5, value=for_empty_passport)
     if ws1_nt == "Thailand":
          ws1_key = f"{str(ws1_fn).strip().lower()}_{str(ws1_ln).strip().lower()}"
          if ws1_key in ws3_fn_ln_ct:
               for_THA_passport = ws3_fn_ln_ct[ws1_key]
               if str(for_THA_passport).isalpha():
                    ws1.cell(row=_, column=6, value=for_THA_passport)
     if ws1_bd is None or str(ws1_bd).strip() == "":
          ws1_key = f"{str(ws1_fn).strip().lower()}_{str(ws1_ln).strip().lower()}"
          if ws1_key in ws3_fn_ln_bd:
               for_empty_birthday = ws3_fn_ln_bd[ws1_key]
               ws1.cell(row=_, column=7, value=sw_date_format(None, None, for_empty_birthday))

for _ in range(ws2.max_row, 4, -1):
     ws2_fn = ws2.cell(row=_, column=10).value
     ws2_ln = ws2.cell(row=_, column=12).value
     ws2_pn = ws2.cell(row=_, column=15).value
     ws2_ct = ws2.cell(row=_, column=13).value
     if ws2_fn and ws2_ln != "":
          ws2_key = f"{str(ws2_fn).strip().lower()}_{str(ws2_ln).strip().lower()}"
          if ws2_key in ws3_fn_ln_pn:
               for_empty_passport = ws3_fn_ln_pn[ws2_key]
               if ws2_pn is None or str(ws2_pn).strip() == "":
                    ws2.cell(row=_, column=15, value=for_empty_passport)
     if ws2_ct == "Thailand":
          ws2_key = f"{str(ws2_fn).strip().lower()}_{str(ws2_ln).strip().lower()}"
          if ws2_key in ws3_fn_ln_ct:
               for_THA_passport = ws3_fn_ln_ct[ws2_key]
               if str(for_THA_passport).isalpha():
                    ws2.cell(row=_, column=13, value=for_THA_passport)
          if ws2_key in ws3_fn_ln_ct:
               for_THA_passport = ws3_fn_ln_ct[ws2_key]
               if str(for_THA_passport).isalpha():
                    ws2.cell(row=_, column=20, value=for_THA_passport)

for _ in range(ws1.max_row, 1, -1):
     nt = ws1.cell(row=_, column=6).value
     if nt in tm_nt_json:
          nt_match = tm_nt_json[nt]
          ws1.cell(row=_, column=6, value=nt_match)

for _ in range(ws2.max_row, 4, -1):
     nt = ws2.cell(row=_, column=13).value
     ct = ws2.cell(row=_, column=20).value
     if nt or ct != "":
          pure_nt = str(nt).strip().lower()
          pure_ct = str(ct).strip().lower()
          for nt, cd in rr_nt_json.items():
               if pure_nt[:5] in str(nt).strip().lower():
                    ws2.cell(row=_, column=13, value=cd)
          for ct, cd in rr_ct_json.items():
               if pure_ct[:5] in str(ct).strip().lower():
                    ws2.cell(row=_, column=21, value=cd)

for _ in range(ws1.max_row, 1, -1):
     pn = ws1.cell(row=_, column=5).value
     nt = ws1.cell(row=_, column=6).value
     if pn is None or pn == "":
          ws1.delete_rows(_, amount=1)
     if nt == "THA":
          ws1.delete_rows(_, amount=1)

for _ in range(ws2.max_row, 4, -1):
     pn = ws2.cell(row=_, column=15).value
     ct = ws2.cell(row=_, column=20).value
     if pn is None or pn == "":
          ws2.delete_rows(_, amount=1)
     if ct is None or pn == "":
          ws2.delete_rows(_, amount=1)

duplicate_pn = set()
for _ in range(ws1.max_row, 1, -1):
    pn = ws1.cell(row=_, column=5).value
    if pn in duplicate_pn:
          ws1.delete_rows(_, amount=1)
    else:
         duplicate_pn.add(pn)

duplicate_pn = set()
for _ in range(ws2.max_row, 4, -1):
    pn = ws2.cell(row=_, column=15).value
    if pn in duplicate_pn:
          ws2.delete_rows(_, amount=1)
    else:
         duplicate_pn.add(pn)

red_color = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for _ in range(ws1.max_row, 1, -1):
     gd = ws1.cell(row=_, column=4).value
     nt = ws1.cell(row=_, column=6).value
     bd = ws1.cell(row=_, column=7).value
     if gd is None or str(gd).strip() == "U" or str(gd).strip() == "":
          ws1.cell(row=_, column=4).fill = red_color
     if nt is None or str(nt).strip() == "":
          ws1.cell(row=_, column=6).fill = red_color
     if len(str(nt).strip()) > 3:
          ws1.cell(row=_, column=6).fill = red_color
     if bd is None or str(bd).strip() == "":
          ws1.cell(row=_, column=7).fill = red_color

for _ in range(ws2.max_row, 4, -1):
     gd = ws2.cell(row=_, column=9).value
     nt = ws2.cell(row=_, column=13).value
     ct = ws2.cell(row=_, column=21).value
     if str(gd).strip() != "":
          if gd == "F":
               ws2.cell(row=_, column=9, value="Ms.")
          elif gd == "M":
               ws2.cell(row=_, column=9, value="Mr.")
          else:
               ws2.cell(row=_, column=9).fill = red_color
     if (str(nt).strip() or str(ct).strip()) is None or (str(nt).strip() or str(ct).strip()) == "":
          ws2.cell(row=_, column=13).fill = red_color
          ws2.cell(row=_, column=21).fill = red_color
     if str(nt).strip().isalpha() or str(ct).strip().isalpha():
          ws2.cell(row=_, column=13).fill = red_color
          ws2.cell(row=_, column=21).fill = red_color

run_num = 1

for _ in range(5, ws2.max_row + 1):
     ws2.cell(row=_, column=1, value=run_num)
     run_num += 1

for _ in range(ws2.max_row, 4, -1):
     ytd = date.today() - timedelta(days=1)
     full_ytd = ytd.strftime("%m/%d/%Y")
     ws2.cell(row=_, column=2, value=sw_date_format(None, full_ytd, None))
     ws2.cell(row=_, column=3, value="15.00")
     ws2.cell(row=_, column=17, value="Phang-nga")
     ws2.cell(row=_, column=18, value="99")
     ws2.cell(row=_, column=19, value="146")
     ws2.cell(row=_, column=22, value="Phuket")
     ws2.cell(row=_, column=23, value="99")
     ws2.cell(row=_, column=25, value="12.00")
     ws2.cell(row=_, column=26, value="1")

ytd = date.today() - timedelta(days=1)
full_ytd = ytd.strftime("%m/%d/%Y")
ws2.cell(row=2, column=14, value=sw_date_format(None, full_ytd, None))

ytd = date.today() - timedelta(days=1)
ytd_y = ytd.strftime("%Y")
ytd_y_be = ytd.year + 543
ytd_m = ytd.strftime("%B")
tm_date = ytd.strftime("%d.%m.%Y")
rr_date = ytd.strftime(f"%d%m{ytd_y_be}")

tm_path = fr"\\LMPC202507256L\Keeper\TM\{ytd_y}\{ytd_m}"
rr_path = fr"\\LMPC202507256L\Keeper\RR.4\{ytd_y}\{ytd_m}"
os.makedirs(tm_path, exist_ok=True)
os.makedirs(rr_path, exist_ok=True)

tm_excel = f"{tm_date}.xlsx"
rr_excel = f"{rr_date}.xlsx"
path_tm_excel = os.path.join(tm_path, tm_excel)
path_rr_excel = os.path.join(rr_path, rr_excel)

wb.save(path_tm_excel)
wb.save(path_rr_excel)

tm_wb = openpyxl.load_workbook(path_tm_excel)
tm_ws = "Sheet1"

for _ in tm_wb.sheetnames:
     if _ != tm_ws:
          del tm_wb[_]

tm_wb.save(path_tm_excel)

rr_wb = openpyxl.load_workbook(path_rr_excel)
rr_ws = "Sheet2"

for _ in rr_wb.sheetnames:
     if _ != rr_ws:
          del rr_wb[_]

rr_ws = rr_wb[rr_ws]
rr_ws.title = "Sheet1"
rr_wb.save(path_rr_excel)

td = date.today()
td_date = td.strftime("%d.%m.%y")

data_path = r"\\LMPC202507256L\Keeper\OTH"
os.makedirs(data_path, exist_ok=True)

new_data_excel = f"get_{td_date}.xlsx"
path_new_data_excel = os.path.join(data_path, new_data_excel)
shutil.copy(path_data_excel, path_new_data_excel)

new_data_wb = load_workbook(path_new_data_excel)

new_data_ws1 = "Sheet1"
new_data_ws2 = "Sheet2"
new_data_ws3 = "Sheet3"

new_data_ws1 = new_data_wb[new_data_ws1]
new_data_ws2 = new_data_wb[new_data_ws2]
new_data_ws3 = new_data_wb[new_data_ws3]

for _ in range(new_data_ws1.max_row, 1, -1):
     new_data_ws1.delete_rows(_, amount=1)
for _ in range(new_data_ws2.max_row, 4, -1):
     new_data_ws2.delete_rows(_, amount=1)
for _ in range(new_data_ws3.max_row, 1, -1):
     new_data_ws3.delete_rows(_, amount=1)

new_data_wb.active = new_data_ws3
new_data_wb.save(path_new_data_excel)