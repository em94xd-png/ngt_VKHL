from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xml.etree.ElementTree
from datetime import datetime, timedelta, date
import json

excel_file = "get_data.xlsx"
ws1 = "Sheet1"
ws2 = "Sheet2"

tree = xml.etree.ElementTree.parse("immigration_report_140258562.XML")
root = tree.getroot()

wb = load_workbook(excel_file)

ws1 = wb[ws1]
ws2 = wb[ws2]

for _ in range(ws1.max_row, 1, -1):
     ws1.delete_rows(_, amount=1)
for _ in range(ws2.max_row, 3, -1):
     ws2.delete_rows(_, amount=1)

with open("tm_ct.json", "r", encoding="utf-8") as file:
     tm_ct_json = json.load(file)
with open("rr_nt.json", "r", encoding="utf-8") as file:
     rr_nt_json = json.load(file)
with open("rr_ct.json", "r", encoding="utf-8") as file:
     rr_ct_json = json.load(file)

def sw_date_format(ad, be):
     if ad and ad.strip() != "":
          try:
               ad = datetime.strptime(ad.strip(), "%m/%d/%Y")
               return ad.strftime("%d/%m/%Y")
          except ValueError:
               return ad
     if be and be.strip() != "":
          try:
               be = datetime.strptime(be.strip(), "%m/%d/%Y")
               cvt_be = be.year + 543
               return be.strftime(f"%d/%m/{cvt_be}")
          except ValueError:
               return be

for _ in root.findall(".//G_IMMIGRATION"):
        fn = _.find("FIRST_NAME").text if _.find("FIRST_NAME") is not None else ""
        ln = _.find("LAST_NAME").text if _.find("LAST_NAME") is not None else ""
        gd = _.find("SEX").text if _.find("SEX") is not None else ""
        pn = _.find("PASSPORT").text if _.find("PASSPORT") is not None else ""
        nt = _.find("NATIONALITY").text if _.find("NATIONALITY") is not None else ""
        bd = _.find("DATE_OF_BIRTH").text if _.find("DATE_OF_BIRTH") is not None else ""
        bd_ad = sw_date_format(bd, None)
        dep = _.find("DEPARTURE_DATE").text if _.find("DEPARTURE_DATE") is not None else ""
        dep_ad = sw_date_format(dep, None)
        dep_be = sw_date_format(None, dep)
        rn = _.find("ROOM").text if _.find("ROOM") is not None else ""
        tm_data = [fn, None, ln, gd, pn, nt, bd_ad, dep_ad]
        rr_data = [None, None, None, rn, None, None, None, None, gd, fn, None, ln, nt, None, pn, None, None, None, None, nt, nt, None, None, dep_be, None, None]
        ws1.append(tm_data)
        ws2.append(rr_data)

ws3 = wb["Sheet3"]
# wb.active = ws3

ws3_fn_ln_pn = {}
ws3_fn_ln_ct = {}
for _ in ws3.iter_rows(values_only=True, min_row=2):
    ws3_fn = _[1]
    ws3_ln = _[0]    
    ws3_pn = _[4]     
    ws3_ct = _[3]
    if ws3_fn and ws3_ln and ws3_pn != "":
          ws3_key = f"{str(ws3_fn).strip()}_{str(ws3_ln).strip()}"
          ws3_fn_ln_pn[ws3_key] = str(ws3_pn).strip()
          ws3_fn_ln_ct[ws3_key] = str(ws3_ct).strip()

for _ in range(ws1.max_row, 1, -1):
    ws1_fn = ws1.cell(row=_, column=1).value
    ws1_ln = ws1.cell(row=_, column=3).value
    ws1_pn = ws1.cell(row=_, column=5).value
    ws1_ct = ws1.cell(row=_, column=6).value
    if ws1_fn and ws1_ln != "":
          ws1_key = f"{str(ws1_fn).strip()}_{str(ws1_ln).strip()}"
          if ws1_key in ws3_fn_ln_pn:
               for_empty_passport = ws3_fn_ln_pn[ws1_key]
               if ws1_pn is None or str(ws1_pn).strip() == "":
                    ws1.cell(row=_, column=5, value=for_empty_passport)
    if ws1_ct == "Thailand":
          ws1_key = f"{str(ws1_fn).strip()}_{str(ws1_ln).strip()}"
          if ws1_key in ws3_fn_ln_ct:
               for_THA_passport = ws3_fn_ln_ct[ws1_key]  
               ws1.cell(row=_, column=6, value=for_THA_passport)

for _ in range(ws2.max_row, 3, -1):
     ws2_fn = ws2.cell(row=_, column=10).value
     ws2_ln = ws2.cell(row=_, column=12).value
     ws2_pn = ws2.cell(row=_, column=15).value
     ws2_ct = ws2.cell(row=_, column=13).value
     if ws2_fn and ws2_ln != "":
          ws2_key = f"{str(ws2_fn).strip()}_{str(ws2_ln).strip()}"
          if ws2_key in ws3_fn_ln_pn:
               for_empty_passport = ws3_fn_ln_pn[ws2_key]
               if ws2_pn is None or str(ws2_pn).strip() == "":
                    ws2.cell(row=_, column=15, value=for_empty_passport)
     if ws2_ct == "Thailand":
          ws2_key = f"{str(ws2_fn).strip()}_{str(ws2_ln).strip()}"
          if ws2_key in ws3_fn_ln_ct:
               for_THA_passport = ws3_fn_ln_ct[ws2_key]
               ws2.cell(row=_, column=13, value=for_THA_passport)

for _ in range(ws1.max_row, 1, -1):
     ct = ws1.cell(row=_, column=6).value
     if ct in tm_ct_json:
          ct_match = tm_ct_json[ct]
          ws1.cell(row=_, column=6, value=ct_match)

for _ in range(ws2.max_row, 3, -1):
     nt = ws2.cell(row=_, column=13).value
     ct = ws2.cell(row=_, column=21).value
     if str(nt).strip() or str(ct).strip() != "":
          pure_nt = str(nt).strip().lower()
          pure_ct = str(ct).strip().lower()
          for nt, cd in rr_nt_json.items():
               if pure_nt in str(nt).lower():
                    ws2.cell(row=_, column=13, value=cd)
          for ct, cd in rr_ct_json.items():
               if pure_ct in str(ct).lower():
                    ws2.cell(row=_, column=21, value=cd)

for _ in range(ws1.max_row, 1, -1):
     pn = ws1.cell(row=_, column=5).value
     if pn is None or pn == "":
          ws1.delete_rows(_, amount=1)

for _ in range(ws2.max_row, 3, -1):
     pn = ws2.cell(row=_, column=15).value
     if pn is None or pn == "":
          ws2.delete_rows(_, amount=1)

for _ in range(ws2.max_row, 3, -1):
     pn = ws2.cell(row=_, column=20).value
     if pn is None or pn == "":
          ws2.delete_rows(_, amount=1)

duplicate_pn = set()
for _ in range(ws1.max_row, 1, -1):
    pn = ws1.cell(row=_, column=5).value
    if pn in duplicate_pn:
          ws1.delete_rows(_, amount=1)
    else:
         duplicate_pn.add(pn)

duplicate_pn = set()
for _ in range(ws2.max_row, 3, -1):
    pn = ws2.cell(row=_, column=15).value
    if pn in duplicate_pn:
          ws2.delete_rows(_, amount=1)
    else:
         duplicate_pn.add(pn)

red_color = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for _ in range(ws2.max_row, 3, -1):
     gd = ws2.cell(row=_, column=9).value
     if str(gd).strip() != "":
          if gd == "F":
               ws2.cell(row=_, column=9, value="Ms.")
          elif gd == "M":
               ws2.cell(row=_, column=9, value="Mr.")
          else:
               ws2.cell(row=_, column=9).fill = red_color

for _ in range(ws2.max_row, 3, -1):
     nt = ws2.cell(row=_, column=13).value
     if not nt or str(nt).strip() == "":
          ws2.cell(row=_, column=13).fill = red_color

run_num = 1

for _ in range(4, ws2.max_row + 1):
     ws2.cell(row=_, column=1, value=run_num)
     run_num += 1

for _ in range(ws2.max_row, 3, -1):
     ytd = date.today() - timedelta(days=1)
     full_ytd = ytd.strftime("%m/%d/%Y")
     ws2.cell(row=_, column=2, value=sw_date_format(None, full_ytd))
     ws2.cell(row=_, column=3, value="15.00")
     ws2.cell(row=_, column=17, value="Phang-nga")
     ws2.cell(row=_, column=18, value="99")
     ws2.cell(row=_, column=19, value="146")
     ws2.cell(row=_, column=22, value="Phuket")
     ws2.cell(row=_, column=23, value="99")
     ws2.cell(row=_, column=25, value="12.00")
     ws2.cell(row=_, column=26, value="1")

ytd = date.today() - timedelta(days=1)
full_ytd = ytd.strftime("%m/%d/%Y")
ws2.cell(row=2, column=14, value=sw_date_format(None, full_ytd))

for _ in wb.worksheets:
     _.views.sheetView[0].tabSelected = False
         
wb.active = wb["Sheet1"]
wb.save(excel_file)