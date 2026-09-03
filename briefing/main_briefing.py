from openpyxl import load_workbook
from datetime import *
from openpyxl.styles import PatternFill
import time, os, xml.etree.ElementTree, sys, pyautogui, subprocess, pygetwindow, pyperclip, win32con, win32gui, json, re

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import script_config

pyautogui.FAILSAFE = True

if not os.path.exists(f"{script_config.path_share}"):
     sys.exit()

# Open Opera
subprocess.run(["cmd", "/c", "start", "msedge", script_config.site_OPERA])

if pygetwindow.getWindowsWithTitle("Opera Cloud"):
     pygetwindow.getWindowsWithTitle("Opera Cloud")[0].restore()
     pygetwindow.getWindowsWithTitle("Opera Cloud")[0].maximize()
time.sleep(.5)

# In Opera
script_config.first_OPERA_open()
script_config.zoom_out(10)
script_config.zoom_in(3)
script_config.main_OPERA_menu()

# To report search
pyautogui.press("tab", presses=5, interval=0.01)
pyautogui.press("right", presses=6, interval=0.01)
pyautogui.press("down", interval=0.01)
pyautogui.press("enter", interval=0.01)
script_config.search_reports()
time.sleep(1)
pyautogui.press("tab", interval=0.01)

# Arrivals
pyautogui.write("Arrivals: Detailed FO", interval=.01)
pyautogui.press("enter", interval=.01)
script_config.search_enter_step1()
script_config.search_enter_step2()
time.sleep(.5)
pyautogui.press("tab", presses=8, interval=.01)
pyautogui.press("down", presses=2, interval=.01)
script_config.search_enter_step3()
pyautogui.press("tab", presses=12, interval=.01)
pyautogui.press("enter", interval=.01)
# Arrivals: Config
script_config.config_report()
time.sleep(1)
pyautogui.hotkey("ctrl", "a", interval=.01)
pyautogui.write(script_config.td_dd_mm, interval=.01)
pyautogui.press("tab", interval=.01)
time.sleep(1)
pyautogui.write(script_config.td_dd_mm, interval=.01)
pyautogui.press("tab", presses=4, interval=.01)
time.sleep(.75)
pyautogui.write(script_config.Room_Type, interval=.01)
pyautogui.press("tab", presses=12, interval=.01)
pyautogui.press("delete", interval=.01)
time.sleep(.5)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("enter", interval=0.01)
# Arrivals: Save
script_config.download_as_2()
pyautogui.click(895, 581, interval=.01)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

pyautogui.press("tab", presses=3, interval=.01)
pyautogui.press("space", interval=.01)
# Arrivals: Download
script_config.download_page()
pyautogui.hotkey("ctrl", "j", interval=.01)
time.sleep(.25)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
script_config.not_download_page()
time.sleep(.5)
pyautogui.hotkey("ctrl", "l", interval=.01)
pyautogui.hotkey("ctrl", "c", interval=.01)
arrival_url = pyperclip.paste()
arrival_file = f"res_detail_{re.search(r"(\d+)\.[xX][mM][lL]", arrival_url).group(1)}.XML"
pyautogui.hotkey("ctrl", "w", interval=.01)

download_page = pygetwindow.getWindowsWithTitle("Untitled")[0]
win32gui.PostMessage(download_page._hWnd, win32con.WM_CLOSE, 0, 0)

arrival = os.path.join(script_config.path_.__add__(r"\Downloads"), arrival_file)

script_config.tab_reserve(1)
pyautogui.press("enter", interval=0.01)
script_config.search_reports()
time.sleep(1)
pyautogui.press("tab", interval=0.01)

# Departures
pyautogui.write("departure_all", interval=.01)
pyautogui.press("enter", interval=0.01)
script_config.search_enter_step1()
script_config.search_enter_step2()
time.sleep(.5)
pyautogui.press("tab", presses=8, interval=0.01)
pyautogui.press("down", presses=2, interval=0.01)
script_config.search_enter_step3()
pyautogui.press("tab", presses=12, interval=0.01)
pyautogui.press("enter", interval=0.01)
# Departures: Config
script_config.config_report()
time.sleep(1)
pyautogui.press("tab", presses=3, interval=0.01)
time.sleep(.75)
pyautogui.write(script_config.Room_Type, interval=.01)
pyautogui.press("tab", interval=0.01)
time.sleep(.5)
pyautogui.click(36, 499, interval=.01) # Pseudo Rooms

while True:
     if pyautogui.pixelMatchesColor(36, 499, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(36, 499, (22, 21, 19), tolerance=10):
          break

pyautogui.click(35, 632, interval=.01) # Membership Type

while True:
     if pyautogui.pixelMatchesColor(35, 632, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(35, 632, (22, 21, 19), tolerance=10):
          break

while True:
     if pyautogui.pixelMatchesColor(280, 631, (255, 255, 255), tolerance=10):
          break

pyautogui.click(280, 631, interval=.01) # Membership Level

while True:
     if pyautogui.pixelMatchesColor(280, 631, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(280, 631, (22, 21, 19), tolerance=10):
          break

while True:
     if pyautogui.pixelMatchesColor(622, 638, (27, 26, 24), tolerance=10):
          break
     
pyautogui.press("tab", presses=21, interval=.01)
pyautogui.press("enter", interval=0.01)
# Departures: Save
script_config.download_as_2()
pyautogui.click(895, 581, interval=.01)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break
     
pyautogui.press("tab", presses=3, interval=.01)
pyautogui.press("space", interval=.01)
# Departures: Download
script_config.download_page()
pyautogui.hotkey("ctrl", "j", interval=.01)
time.sleep(.25)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
script_config.not_download_page()
time.sleep(.5)
pyautogui.hotkey("ctrl", "l", interval=.01)
pyautogui.hotkey("ctrl", "c", interval=.01)
departure_url = pyperclip.paste()
departure_file = f"departure_all_{re.search(r"(\d+)\.[xX][mM][lL]", departure_url).group(1)}.XML"
pyautogui.hotkey("ctrl", "w", interval=.01)

download_page = pygetwindow.getWindowsWithTitle("Untitled")[0]
win32gui.PostMessage(download_page._hWnd, win32con.WM_CLOSE, 0, 0)

departure = os.path.join(script_config.path_.__add__(r"\Downloads"), departure_file)

script_config.tab_reserve(1)
pyautogui.press("enter", interval=0.01)
script_config.search_reports()
time.sleep(1)
pyautogui.press("tab", interval=0.01)

# History and Forecast
pyautogui.write("History and Forecast FO", interval=.01)
pyautogui.press("enter", interval=0.01)
script_config.search_enter_step1()
script_config.search_enter_step2()
time.sleep(.5)
pyautogui.press("tab", presses=8, interval=0.01)
pyautogui.press("down", presses=2, interval=0.01)
script_config.search_enter_step3()
pyautogui.press("tab", presses=12, interval=0.01)
pyautogui.press("enter", interval=0.01)
# History and Forecast: Config
script_config.config_report()
time.sleep(1)
pyautogui.hotkey("ctrl", "a", interval=.01)
pyautogui.write(script_config.ytd_dd_mm, interval=.01)
pyautogui.press("tab", interval=.01)
time.sleep(1)
pyautogui.write("+5", interval=.01)
pyautogui.press("tab", presses=7, interval=0.01)
pyautogui.press("enter", interval=0.01)
# History and Forecast: Save
script_config.download_as_2()
pyautogui.click(895, 581, interval=.01)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

pyautogui.press("tab", presses=3, interval=.01)
pyautogui.press("space", interval=.01)
# History and Forecast: Download
script_config.download_page()
pyautogui.hotkey("ctrl", "j", interval=.01)
time.sleep(.25)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
script_config.not_download_page()
time.sleep(.5)
pyautogui.hotkey("ctrl", "l", interval=.01)
pyautogui.hotkey("ctrl", "c", interval=.01)
history_forecast_url = pyperclip.paste()
history_forecast_file = f"history_forecast_{re.search(r"(\d+)\.[xX][mM][lL]", history_forecast_url).group(1)}.XML"
pyautogui.hotkey("ctrl", "w", interval=.01)

download_page = pygetwindow.getWindowsWithTitle("Untitled")[0]
win32gui.PostMessage(download_page._hWnd, win32con.WM_CLOSE, 0, 0)

history_forecast = os.path.join(script_config.path_.__add__(r"\Downloads"), history_forecast_file)

# History and Forecast (AVC)
# History and Forecast (AVC): Config
time.sleep(.5)
pyautogui.click(512, 445, interval=.01) # Pseudo Rooms

while True:
     if pyautogui.pixelMatchesColor(512, 445, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(512, 445, (22, 21, 19), tolerance=10):
          break

pyautogui.press("tab", presses=9, interval=.01)
pyautogui.press("enter", interval=0.01)
# History and Forecast (AVC): Save
script_config.download_as_2()
pyautogui.click(895, 581, interval=.01)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

pyautogui.press("tab", presses=3, interval=.01)
pyautogui.press("space", interval=.01)
# History and Forecast (AVC): Download
script_config.download_page()
pyautogui.hotkey("ctrl", "j", interval=.01)
time.sleep(.25)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
script_config.not_download_page()
time.sleep(.5)
pyautogui.hotkey("ctrl", "l", interval=.01)
pyautogui.hotkey("ctrl", "c", interval=.01)
history_forecast_AVC_url = pyperclip.paste()
history_forecast_AVC_file = f"history_forecast_{re.search(r"(\d+)\.[xX][mM][lL]", history_forecast_AVC_url).group(1)}.XML"
pyautogui.hotkey("ctrl", "w", interval=.01)

download_page = pygetwindow.getWindowsWithTitle("Untitled")[0]
win32gui.PostMessage(download_page._hWnd, win32con.WM_CLOSE, 0, 0)

history_forecast_AVC = os.path.join(script_config.path_.__add__(r"\Downloads"), history_forecast_AVC_file)

script_config.tab_reserve(1)
pyautogui.press("enter", interval=0.01)
script_config.search_reports()
time.sleep(1)
pyautogui.press("tab", interval=0.01)

# Forecast
pyautogui.write("resfutureoccupancy", interval=.01)
pyautogui.press("enter", interval=0.01)
script_config.search_enter_step1()
script_config.search_enter_step2()
time.sleep(.5)
pyautogui.press("tab", presses=8, interval=0.01)
pyautogui.press("down", presses=2, interval=0.01)
script_config.search_enter_step3()
pyautogui.press("tab", presses=12, interval=0.01)
pyautogui.press("enter", interval=0.01)
# Forecast: Config
script_config.config_report()
time.sleep(1)
pyautogui.hotkey("ctrl", "a", interval=.01)
pyautogui.write(script_config.td_dd_mm, interval=.01)
pyautogui.press("tab", interval=.01)
time.sleep(1)
pyautogui.write(script_config.td_dd_mm, interval=.01)
pyautogui.press("tab", presses=4, interval=.01)
pyautogui.press("space", interval=0.01)
# Forecast: Save
script_config.download_as_2()
pyautogui.click(895, 581, interval=.01)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

pyautogui.press("tab", presses=3, interval=.01)
pyautogui.press("space", interval=.01)
# Forecast: Download
script_config.download_page()
pyautogui.hotkey("ctrl", "j", interval=.01)
time.sleep(.25)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
script_config.not_download_page()
time.sleep(.5)
pyautogui.hotkey("ctrl", "l", interval=.01)
pyautogui.hotkey("ctrl", "c", interval=.01)
resfutureoccupancy_url = pyperclip.paste()
resfutureoccupancy_file = f"resfutureoccupancy_{re.search(r"(\d+)\.[xX][mM][lL]", resfutureoccupancy_url).group(1)}.XML"
pyautogui.hotkey("ctrl", "w", interval=.01)

download_page = pygetwindow.getWindowsWithTitle("Untitled")[0]
win32gui.PostMessage(download_page._hWnd, win32con.WM_CLOSE, 0, 0)

resfutureoccupancy = os.path.join(script_config.path_.__add__(r"\Downloads"), resfutureoccupancy_file)

script_config.tab_reserve(1)
pyautogui.press("enter", interval=0.01)
script_config.search_reports()
time.sleep(1)
pyautogui.press("tab", interval=0.01)

# Room Upgrade
pyautogui.write("Journal by Cashier and Transaction Code", interval=.01)
pyautogui.press("enter", interval=0.01)
script_config.search_enter_step1()
script_config.search_enter_step2()
time.sleep(.5)
pyautogui.press("tab", presses=8, interval=0.01)
pyautogui.press("down", presses=2, interval=0.01)
script_config.search_enter_step3()
pyautogui.press("tab", presses=12, interval=0.01)
pyautogui.press("enter", interval=0.01)
# Room Upgrade: Config
script_config.config_report()
time.sleep(1)
pyautogui.hotkey("ctrl", "a", interval=.01)
pyautogui.write(script_config.ytd_dd_mm, interval=.01)
pyautogui.press("tab", interval=0.01)
time.sleep(1)
pyautogui.write(script_config.ytd_dd_mm, interval=.01)
pyautogui.press("tab", presses=4, interval=0.01)
time.sleep(1)
pyautogui.write("10203", interval=.01)
pyautogui.press("tab", presses=8, interval=0.01)
pyautogui.press("space", interval=.01)
# Room Upgrade: Save
script_config.download_as_2()
pyautogui.click(895, 581, interval=.01)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

pyautogui.press("tab", presses=3, interval=.01)
pyautogui.press("space", interval=.01)
# Room Upgrade: Download
script_config.download_page()
pyautogui.hotkey("ctrl", "j", interval=.01)
time.sleep(.25)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
script_config.not_download_page()
time.sleep(.5)
pyautogui.hotkey("ctrl", "l", interval=.01)
pyautogui.hotkey("ctrl", "c", interval=.01)
room_upgrade_url = pyperclip.paste()
room_upgrade_file = f"finjrnlbytrans_{re.search(r"(\d+)\.[xX][mM][lL]", room_upgrade_url).group(1)}.XML"
pyautogui.hotkey("ctrl", "w", interval=.01)

download_page = pygetwindow.getWindowsWithTitle("Untitled")[0]
win32gui.PostMessage(download_page._hWnd, win32con.WM_CLOSE, 0, 0)

room_upgrade = os.path.join(script_config.path_.__add__(r"\Downloads"), room_upgrade_file)

# Late Checkout: Config
script_config.tab_reserve(9)
time.sleep(1)
pyautogui.write("10200,10400", interval=.01)
pyautogui.press("tab", presses=9, interval=.01)
pyautogui.press("space", interval=0.01)
# Late Checkout: Save
script_config.download_as_2()
pyautogui.click(895, 581, interval=.01)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

pyautogui.press("tab", presses=3, interval=.01)
pyautogui.press("space", interval=.01)
# Late Checkout: Download
script_config.download_page()
pyautogui.hotkey("ctrl", "j", interval=.01)
time.sleep(.25)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
script_config.not_download_page()
time.sleep(.5)
pyautogui.hotkey("ctrl", "l", interval=.01)
pyautogui.hotkey("ctrl", "c", interval=.01)
late_checkout_url = pyperclip.paste()
late_checkout_file = f"finjrnlbytrans_{re.search(r"(\d+)\.[xX][mM][lL]", late_checkout_url).group(1)}.XML"
pyautogui.hotkey("ctrl", "w", interval=.01)

download_page = pygetwindow.getWindowsWithTitle("Untitled")[0]
win32gui.PostMessage(download_page._hWnd, win32con.WM_CLOSE, 0, 0)

late_checkout = os.path.join(script_config.path_.__add__(r"\Downloads"), late_checkout_file)

# Tour Commission: Config
script_config.tab_reserve(9)
time.sleep(1)
pyautogui.write("62710", interval=.01)
pyautogui.press("tab", presses=9, interval=.01)
pyautogui.press("space", interval=0.01)
# Tour Commission: Save
script_config.download_as_2()
pyautogui.click(895, 581, interval=.01)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

pyautogui.press("tab", presses=3, interval=.01)
pyautogui.press("space", interval=.01)
# Tour Commission: Download
script_config.download_page()
pyautogui.hotkey("ctrl", "j", interval=.01)
time.sleep(.25)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
script_config.not_download_page()
time.sleep(.5)
pyautogui.hotkey("ctrl", "l", interval=.01)
pyautogui.hotkey("ctrl", "c", interval=.01)
tour_commission_url = pyperclip.paste()
tour_commission_file = f"finjrnlbytrans_{re.search(r"(\d+)\.[xX][mM][lL]", tour_commission_url).group(1)}.XML"
pyautogui.hotkey("ctrl", "w", interval=.01)

download_page = pygetwindow.getWindowsWithTitle("Untitled")[0]
win32gui.PostMessage(download_page._hWnd, win32con.WM_CLOSE, 0, 0)

tour_commission = os.path.join(script_config.path_.__add__(r"\Downloads"), tour_commission_file)

# Gift Shop: Config
script_config.tab_reserve(9)
time.sleep(1)
pyautogui.write("60600,60601,60630,60631", interval=.01)
pyautogui.press("tab", presses=9, interval=.01)
pyautogui.press("enter", interval=0.01)
# Gift Shop: Save
script_config.download_as_2()
pyautogui.click(895, 581, interval=.01)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

pyautogui.press("tab", presses=3, interval=.01)
pyautogui.press("space", interval=.01)
# Gift Shop: Download
script_config.download_page()
pyautogui.hotkey("ctrl", "j", interval=.01)
time.sleep(.25)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
script_config.not_download_page()
time.sleep(.5)
pyautogui.hotkey("ctrl", "l", interval=.01)
pyautogui.hotkey("ctrl", "c", interval=.01)
gift_shop_url = pyperclip.paste()
gift_shop_file = f"finjrnlbytrans_{re.search(r"(\d+)\.[xX][mM][lL]", gift_shop_url).group(1)}.XML"
pyautogui.hotkey("ctrl", "w", interval=.01)

download_page = pygetwindow.getWindowsWithTitle("Untitled")[0]
win32gui.PostMessage(download_page._hWnd, win32con.WM_CLOSE, 0, 0)

gift_shop = os.path.join(script_config.path_.__add__(r"\Downloads"), gift_shop_file)

# Room Upgrade MTD: Config
script_config.tab_reserve(14)
time.sleep(1)
pyautogui.write(f"{script_config.ytd.replace(day=1).strftime("%d%m")}", interval=.01)
pyautogui.press("tab", interval=.01)
time.sleep(1)
pyautogui.write(f"{script_config.ytd_dd_mm}", interval=.01)
pyautogui.press("tab", presses=4, interval=0.01)
time.sleep(1)
pyautogui.write("10203", interval=.01)
pyautogui.press("tab", presses=9, interval=0.01)
pyautogui.press("space", interval=.01)
# Room Upgrade MTD: Save
script_config.download_as_2()
pyautogui.click(895, 581, interval=.01)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

pyautogui.press("tab", presses=3, interval=.01)
pyautogui.press("space", interval=.01)
# Room Upgrade MTD: Download
script_config.download_page()
pyautogui.hotkey("ctrl", "j", interval=.01)
time.sleep(.25)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
script_config.not_download_page()
time.sleep(.5)
pyautogui.hotkey("ctrl", "l", interval=.01)
pyautogui.hotkey("ctrl", "c", interval=.01)
room_upgrade_mtd_url = pyperclip.paste()
room_upgrade_mtd_file = f"finjrnlbytrans_{re.search(r"(\d+)\.[xX][mM][lL]", room_upgrade_mtd_url).group(1)}.XML"
pyautogui.hotkey("ctrl", "w", interval=.01)

download_page = pygetwindow.getWindowsWithTitle("Untitled")[0]
win32gui.PostMessage(download_page._hWnd, win32con.WM_CLOSE, 0, 0)

room_upgrade_mtd = os.path.join(script_config.path_.__add__(r"\Downloads"), room_upgrade_mtd_file)

# Late Checkout MTD: Config
script_config.tab_reserve(9)
time.sleep(1)
pyautogui.write("10200,10400", interval=.01)
pyautogui.press("tab", presses=9, interval=.01)
pyautogui.press("space", interval=0.01)
# Late Checkout MTD: Save
script_config.download_as_2()
pyautogui.click(895, 581, interval=.01)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

pyautogui.press("tab", presses=3, interval=.01)
pyautogui.press("space", interval=.01)
# Late Checkout MTD: Download
script_config.download_page()
pyautogui.hotkey("ctrl", "j", interval=.01)
time.sleep(.25)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
script_config.not_download_page()
time.sleep(.5)
pyautogui.hotkey("ctrl", "l", interval=.01)
pyautogui.hotkey("ctrl", "c", interval=.01)
late_checkout_mtd_url = pyperclip.paste()
late_checkout_mtd_file = f"finjrnlbytrans_{re.search(r"(\d+)\.[xX][mM][lL]", late_checkout_mtd_url).group(1)}.XML"
pyautogui.hotkey("ctrl", "w", interval=.01)

download_page = pygetwindow.getWindowsWithTitle("Untitled")[0]
win32gui.PostMessage(download_page._hWnd, win32con.WM_CLOSE, 0, 0)

late_checkout_mtd = os.path.join(script_config.path_.__add__(r"\Downloads"), late_checkout_mtd_file)

# Tour Commission MTD: Config
script_config.tab_reserve(9)
time.sleep(1)
pyautogui.write("62710", interval=.01)
pyautogui.press("tab", presses=9, interval=.01)
pyautogui.press("enter", interval=0.01)
# Tour Commission MTD: Save
script_config.download_as_2()
pyautogui.click(895, 581, interval=.01)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

pyautogui.press("tab", presses=3, interval=.01)
pyautogui.press("space", interval=.01)
# Tour Commission MTD: Download
script_config.download_page()
pyautogui.hotkey("ctrl", "j", interval=.01)
time.sleep(.25)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
script_config.not_download_page()
time.sleep(.5)
pyautogui.hotkey("ctrl", "l", interval=.01)
pyautogui.hotkey("ctrl", "c", interval=.01)
tour_commission_mtd_url = pyperclip.paste()
tour_commission_mtd_file = f"finjrnlbytrans_{re.search(r"(\d+)\.[xX][mM][lL]", tour_commission_mtd_url).group(1)}.XML"
pyautogui.hotkey("ctrl", "w", interval=.01)

download_page = pygetwindow.getWindowsWithTitle("Untitled")[0]
win32gui.PostMessage(download_page._hWnd, win32con.WM_CLOSE, 0, 0)

tour_commission_mtd = os.path.join(script_config.path_.__add__(r"\Downloads"), tour_commission_mtd_file)

# Gift Shop MTD: Config
script_config.tab_reserve(9)
time.sleep(1)
pyautogui.write("60600,60601,60630,60631", interval=.01)
pyautogui.press("tab", presses=9, interval=.01)
pyautogui.press("enter", interval=0.01)
# Gift Shop MTD: Save
script_config.download_as_2()
pyautogui.click(895, 581, interval=.01)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(895, 581, (22, 21, 19), tolerance=10):
          break

pyautogui.press("tab", presses=3, interval=.01)
pyautogui.press("space", interval=.01)
# Gift Shop MTD: Download
script_config.download_page()
pyautogui.hotkey("ctrl", "j", interval=.01)
time.sleep(.25)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
script_config.not_download_page()
time.sleep(.5)
pyautogui.hotkey("ctrl", "l", interval=.01)
pyautogui.hotkey("ctrl", "c", interval=.01)
gift_shop_mtd_url = pyperclip.paste()
gift_shop_mtd_file = f"finjrnlbytrans_{re.search(r"(\d+)\.[xX][mM][lL]", gift_shop_mtd_url).group(1)}.XML"
pyautogui.hotkey("ctrl", "w", interval=.01)

download_page = pygetwindow.getWindowsWithTitle("Untitled")[0]
win32gui.PostMessage(download_page._hWnd, win32con.WM_CLOSE, 0, 0)

gift_shop_mtd = os.path.join(script_config.path_.__add__(r"\Downloads"), gift_shop_mtd_file)

script_config.tab_reserve(1)
pyautogui.press("enter", interval=0.01)
script_config.search_reports()
time.sleep(1)
pyautogui.press("tab", interval=0.01)

# ARR Immigration
pyautogui.write("immigration_report", interval=.01)
pyautogui.press("enter", interval=0.01)
script_config.search_enter_step1()
script_config.search_enter_step2()
time.sleep(.5)
pyautogui.press("tab", presses=8, interval=0.01)
pyautogui.press("down", presses=2, interval=0.01)
script_config.search_enter_step3()
pyautogui.press("tab", presses=12, interval=0.01)
pyautogui.press("enter", interval=0.01)
# ARR Immigration: Config
script_config.config_report()
time.sleep(1)
pyautogui.click(35, 375, interval=.01)

while True:
     if pyautogui.pixelMatchesColor(35, 375, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(35, 375, (22, 21, 19), tolerance=10):
          break

pyautogui.press("tab", presses=2, interval=.01)
pyautogui.press("space", interval=.01) # Download As...
# ARR Immigration: Save
script_config.download_as()
pyautogui.click(896, 596, interval=.01)

while True:
     if pyautogui.pixelMatchesColor(896, 596, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(896, 596, (22, 21, 19), tolerance=10):
          break

pyautogui.press("tab", presses=3, interval=.01)
pyautogui.press("space", interval=.01)
# ARR Immigration: Download
script_config.download_page()
pyautogui.hotkey("ctrl", "j", interval=.01)
time.sleep(.25)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
script_config.not_download_page()
time.sleep(.5)
pyautogui.hotkey("ctrl", "l", interval=.01)
pyautogui.hotkey("ctrl", "c", interval=.01)
arr_immigration_url = pyperclip.paste()
arr_immigration_file = f"immigration_report_{re.search(r"(\d+)\.[xX][mM][lL]", arr_immigration_url).group(1)}.XML"
pyautogui.hotkey("ctrl", "w", interval=.01)

download_page = pygetwindow.getWindowsWithTitle("Untitled")[0]
win32gui.PostMessage(download_page._hWnd, win32con.WM_CLOSE, 0, 0)

arr_immigration = os.path.join(script_config.path_.__add__(r"\Downloads"), arr_immigration_file)

# DEP Immigration: Config
script_config.tab_reserve(3)
time.sleep(1)
pyautogui.write("CHECKED OUT,DEPARTURE", interval=.01)
pyautogui.press("tab", presses=3, interval=.01)
pyautogui.press("space", interval=.01)
# DEP Immigration: Save
script_config.download_as()
pyautogui.click(896, 596, interval=.01)

while True:
     if pyautogui.pixelMatchesColor(896, 596, (22, 21, 19), tolerance=10):
          break

time.sleep(.5)

while True:
     if pyautogui.pixelMatchesColor(896, 596, (22, 21, 19), tolerance=10):
          break

pyautogui.press("tab", presses=3, interval=.01)
pyautogui.press("space", interval=.01)
# DEP Immigration: Download
script_config.download_page()
pyautogui.hotkey("ctrl", "j", interval=.01)
time.sleep(.25)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.press("tab", presses=6, interval=.01)
pyautogui.press("space", interval=.01)
script_config.not_download_page()
time.sleep(.5)
pyautogui.hotkey("ctrl", "l", interval=.01)
pyautogui.hotkey("ctrl", "c", interval=.01)
dep_immigration_url = pyperclip.paste()
dep_immigration_file = f"immigration_report_{re.search(r"(\d+)\.[xX][mM][lL]", dep_immigration_url).group(1)}.XML"
pyautogui.hotkey("ctrl", "w", interval=.01)

download_page = pygetwindow.getWindowsWithTitle("Untitled")[0]
win32gui.PostMessage(download_page._hWnd, win32con.WM_CLOSE, 0, 0)

dep_immigration = os.path.join(script_config.path_.__add__(r"\Downloads"), dep_immigration_file)

path_OTH = script_config.path_share.__add__(r"\OTH")

ori_excel_file = "briefing.xlsm"
path_ori_excel = os.path.join(path_OTH, ori_excel_file)

path_td_excel = os.path.join(path_OTH, f"{script_config.td_dot_dd_mm_yy}.xlsm")

wb = load_workbook(path_ori_excel, keep_vba=True)
wb.save(path_td_excel)

os.startfile(path_td_excel)
time.sleep(2.5)

if pygetwindow.getWindowsWithTitle("Excel"):
    pygetwindow.getWindowsWithTitle("Excel")[0].activate()
    pygetwindow.getWindowsWithTitle("Excel")[0].maximize()
if not pygetwindow.getWindowsWithTitle("Excel"):
     sys.exit()

script_config.stay_excel()

pyautogui.hotkey("alt", "f11", interval=.01)
pyautogui.press("alt", interval=.01)
pyautogui.press("i", interval=.01)
pyautogui.press("m", interval=.01)

tree = xml.etree.ElementTree.parse(arrival)
root = tree.getroot()

arr = root.findall(".//G_RESERVATION")
arr_row = 1 + len(arr)

tree = xml.etree.ElementTree.parse(departure)
root = tree.getroot()

dep = root.findall(".//G_ROOM")
dep_row = 1 + len(dep)

set_row = f"""Sub Macro1()
'
' Macro1 Macro
'

'
    ActiveWindow.SmallScroll Down:=0
    Rows("16:16").Select
{script_config.num_row(arr_row - 2)}

    ActiveWindow.SmallScroll Down:=2
    Rows("{arr_row + 17}:{arr_row + 17}").Select
{script_config.num_row(dep_row - 2)}
    
End Sub
"""

pyperclip.copy(set_row)
time.sleep(.25)
pyautogui.hotkey("ctrl", "v", interval=.01)
pyautogui.press("f5", interval=.01)
pyautogui.hotkey("alt", "f4")

script_config.stay_excel()

pyautogui.hotkey("ctrl", "s")
pyautogui.hotkey("alt", "f4")

script_config.not_stay_excel()

wb = load_workbook(path_td_excel, keep_vba=True)

ws = "Ori Format"
ws = wb[ws]
ws.title = f"{script_config.td_dot_dd_mm_yy}"

ws["A2"] = f"Daily Briefing\n{script_config.td_full_d} {script_config.td_full_m} {script_config.td_short_date}, {script_config.td_yyyy}"

tree = xml.etree.ElementTree.parse(history_forecast)
root = tree.getroot()

for _ in root.findall(".//G_GPAGEID/LIST_G_REC_TYPE/G_REC_TYPE/LIST_G_CONSIDERED_DATE/G_CONSIDERED_DATE"):
    hu = _.find("HOUSE_USE_ROOMS").text if _.find("HOUSE_USE_ROOMS").text is not None else "N/A"
    arr = _.find("ARRIVAL_ROOMS").text if _.find("ARRIVAL_ROOMS").text is not None else "-"
    dep = _.find("DEPARTURE_ROOMS").text if _.find("DEPARTURE_ROOMS").text is not None else "-"
    ppl = _.find("NO_PERSONS").text if _.find("NO_PERSONS").text is not None else "-"
    if _.find("CONSIDERED_DATE").text == f"{script_config.td_hp_dd_mm_yy.upper()}":
        ws["E5"] = arr
        ws["E7"] = dep
    if _.find("CONSIDERED_DATE").text == f"{(date.today() + timedelta(days=1)).strftime("%d-%b-%y").upper()}":
        ws["M5"] = arr
        ws["M7"] = dep
    if _.find("CONSIDERED_DATE").text == f"{(date.today() + timedelta(days=2)).strftime("%d-%b-%y").upper()}":
        ws["O5"] = arr
        ws["O7"] = dep
    if _.find("CONSIDERED_DATE").text == f"{(date.today() + timedelta(days=3)).strftime("%d-%b-%y").upper()}":
        ws["Q5"] = arr
        ws["Q7"] = dep
    if _.find("CONSIDERED_DATE").text == f"{(date.today() + timedelta(days=4)).strftime("%d-%b-%y").upper()}":
        ws["S5"] = arr
        ws["S7"] = dep
    if _.find("CONSIDERED_DATE").text == f"{(date.today() + timedelta(days=5)).strftime("%d-%b-%y").upper()}":
        ws["U5"] = arr
        ws["U7"] = dep

tree = xml.etree.ElementTree.parse(history_forecast_AVC)
root = tree.getroot()

for _ in root.findall(".//G_GPAGEID/LIST_G_REC_TYPE/G_REC_TYPE/LIST_G_CONSIDERED_DATE/G_CONSIDERED_DATE"):
    cmp = _.find("COMPLIMENTARY_ROOMS").text if _.find("COMPLIMENTARY_ROOMS").text is not None else "-"
    if cmp == "0":
        cmp = "-"
    hu = int(_.find("HOUSE_USE_ROOMS").text if str(_.find("HOUSE_USE_ROOMS").text) else 0)
    if hu == "0":
        hu = "-"
    ns = _.find("NO_SHOW_ROOMS").text if _.find("NO_SHOW_ROOMS").text else "-"
    if ns == "0":
        ns = "-"
    adr = round(float(_.find("CF_AVERAGE_ROOM_RATE").text), 2) if _.find("CF_AVERAGE_ROOM_RATE").text is not None else "-"
    arr = _.find("ARRIVAL_ROOMS").text if _.find("ARRIVAL_ROOMS").text is not None else "-"
    dep = _.find("DEPARTURE_ROOMS").text if _.find("DEPARTURE_ROOMS").text is not None else "-"
    rm = int(_.find("NO_ROOMS").text if str(_.find("NO_ROOMS").text) else 0)
    ppl = _.find("NO_PERSONS").text if _.find("NO_PERSONS").text is not None else "-"
    oc = round(float(_.find("CF_OCCUPANCY").text), 2) if _.find("CF_OCCUPANCY").text is not None else "-"
    if _.find("CONSIDERED_DATE").text == f"{script_config.ytd_hp_dd_mm_yy.upper()}":
        ws["B5"] = cmp
        ws["B7"] = hu
        ws["B9"] = ns
        ws["B11"] = adr
        ws["B11"].number_format = "#,###.##"
        ws["B12"] = oc
        if (ws["B12"].value).is_integer():
            ws["B12"].number_format = "#\"%\""
        elif round(float(ws["B12"].value), 2):
            ws["B12"].number_format = "#.##\"%\""
    if _.find("CONSIDERED_DATE").text == f"{script_config.td_hp_dd_mm_yy.upper()}":
        ws["D4"] = f"{script_config.td.strftime("%#d-%b")}"
        ws["F5"] = arr
        ws["F7"] = dep
        ws["E9"] = rm
        ws["G12"] = ppl
    if _.find("CONSIDERED_DATE").text == f"{(date.today() + timedelta(days=1)).strftime("%d-%b-%y").upper()}":
        ws["N5"] = arr
        ws["N7"] = dep
        ws["M9"] = rm
        ws["M11"] = (rm - hu) / (327 - hu) * 100
        if (ws["M11"].value).is_integer():
            ws["M11"].number_format = "#\"%\""
        elif round(float(ws["M11"].value), 2):
            ws["M11"].number_format = "#.##\"%\""
        ws["M12"] = ppl
    if _.find("CONSIDERED_DATE").text == f"{(date.today() + timedelta(days=2)).strftime("%d-%b-%y").upper()}":
        ws["P5"] = arr
        ws["P7"] = dep
        ws["O9"] = rm
        ws["O11"] = (rm - hu) / (327 - hu) * 100
        if (ws["O11"].value).is_integer():
            ws["O11"].number_format = "#\"%\""
        elif round(float(ws["O11"].value), 2):
            ws["O11"].number_format = "#.##\"%\""
        ws["O12"] = ppl
    if _.find("CONSIDERED_DATE").text == f"{(date.today() + timedelta(days=3)).strftime("%d-%b-%y").upper()}":
        ws["R5"] = arr
        ws["R7"] = dep
        ws["Q9"] = rm
        ws["Q11"] = (rm - hu) / (327 - hu) * 100
        if (ws["Q11"].value).is_integer():
            ws["Q11"].number_format = "#\"%\""
        elif round(float(ws["Q11"].value), 2):
            ws["Q11"].number_format = "#.##\"%\""
        ws["Q12"] = ppl
    if _.find("CONSIDERED_DATE").text == f"{(date.today() + timedelta(days=4)).strftime("%d-%b-%y").upper()}":
        ws["T5"] = arr
        ws["T7"] = dep
        ws["S9"] = rm
        ws["S11"] = (rm - hu) / (327 - hu) * 100
        if (ws["S11"].value).is_integer():
            ws["S11"].number_format = "#\"%\""
        elif round(float(ws["S11"].value), 2):
            ws["S11"].number_format = "#.##\"%\""
        ws["S12"] = ppl
    if _.find("CONSIDERED_DATE").text == f"{(date.today() + timedelta(days=5)).strftime("%d-%b-%y").upper()}":
        ws["V5"] = arr
        ws["V7"] = dep
        ws["U9"] = rm
        ws["U11"] = (rm - hu) / (327 - hu) * 100
        if (ws["U11"].value).is_integer():
            ws["U11"].number_format = "#\"%\""
        elif round(float(ws["U11"].value), 2):
            ws["U11"].number_format = "#.##\"%\""
        ws["U12"] = ppl

cmp = int(ws["B5"].value or 0 if ws["B5"].value != "-" else 0)
hu = int(ws["B7"].value or 0 if ws["B7"].value != "-" else 0)
rm = int(ws["E9"].value or 0 if ws["E9"].value != "-" else 0)

ws["E11"] = round(float((rm - (cmp + hu)) / 327 * 100), 2)
if (ws["E11"].value).is_integer():
    ws["E11"].number_format = "#\"%\""
elif round(float(ws["E11"].value), 2):
    ws["E11"].number_format = "#.##\"%\""

tree = xml.etree.ElementTree.parse(resfutureoccupancy)
root = tree.getroot()

for _ in root.findall(".//G_RESV_TYPE"):
    adl = _.find("SUMADULTS").text if _.find("SUMADULTS").text is not None else "-"
    chd = _.find("SUMCHILDREN").text if _.find("SUMCHILDREN").text is not None else "-"
    if _.find(".//D_DATE").text == f"{script_config.td_hp_dd_mm_yy.upper()}":
        ws["E12"] = adl
        ws["E13"] = chd

tree = xml.etree.ElementTree.parse(room_upgrade)
root = tree.getroot()

for _ in root:
    if _.tag == "R_DEBIT":
        val = float(_.text)
        if _.text == "0":
            ws["I5"] = "-"
        else:
            if val.is_integer():
                ws["I5"] = int(val)
                ws["I5"].number_format = "#,###"
            else:
                ws["I5"] = round(val, 2)
                ws["I5"].number_format = "#,###.##"

tree = xml.etree.ElementTree.parse(room_upgrade_mtd)
root = tree.getroot()

for _ in root:
    if _.tag == "R_DEBIT":
        val = float(_.text)
        if _.text == "0":
            ws["J5"] = "-"
        else:
            if val.is_integer():
                ws["J5"] = int(val)
                ws["J5"].number_format = "#,###"
            else:
                ws["J5"] = round(val, 2)
                ws["J5"].number_format = "#,###.##"

tree = xml.etree.ElementTree.parse(late_checkout)
root = tree.getroot()

for _ in root:
    if _.tag == "R_DEBIT":
        val = float(_.text)
        if _.text == "0":
            ws["I7"] = "-"
        else:
            if val.is_integer():
                ws["I7"] = int(val)
                ws["I7"].number_format = "#,###"
            else:
                ws["I7"] = round(val, 2)
                ws["I7"].number_format = "#,###.##"

tree = xml.etree.ElementTree.parse(late_checkout_mtd)
root = tree.getroot()

for _ in root:
    if _.tag == "R_DEBIT":
        val = float(_.text)
        if _.text == "0":
            ws["J7"] = "-"
        else:
            if val.is_integer():
                ws["J7"] = int(val)
                ws["J7"].number_format = "#,###"
            else:
                ws["J7"] = round(val, 2)
                ws["J7"].number_format = "#,###.##"

tree = xml.etree.ElementTree.parse(tour_commission)
root = tree.getroot()

for _ in root:
    if _.tag == "R_DEBIT":
        val = float(_.text)
        if _.text == "0":
            ws["I9"] = "-"
        else:
            if val.is_integer():
                ws["I9"] = int(val)
                ws["I9"].number_format = "#,###"
            else:
                ws["I9"] = round(val, 2)
                ws["I9"].number_format = "#,###.##"

tree = xml.etree.ElementTree.parse(tour_commission_mtd)
root = tree.getroot()

for _ in root:
    if _.tag == "R_DEBIT":
        val = float(_.text)
        if _.text == "0":
            ws["J9"] = "-"
        else:
            if val.is_integer():
                ws["J9"] = int(val)
                ws["J9"].number_format = "#,###"
            else:
                ws["J9"] = round(val, 2)
                ws["J9"].number_format = "#,###.##"

tree = xml.etree.ElementTree.parse(gift_shop)
root = tree.getroot()

for _ in root:
    if _.tag == "R_DEBIT":
        val = float(_.text)
        if _.text == "0":
            ws["I12"] = "-"
        else:
            if val.is_integer():
                ws["I12"] = int(val)
                ws["I12"].number_format = "#,###"
            else:
                ws["I12"] = round(val, 2)
                ws["I12"].number_format = "#,###.##"

tree = xml.etree.ElementTree.parse(gift_shop_mtd)
root = tree.getroot()

for _ in root:
    if _.tag == "R_DEBIT":
        val = float(_.text)
        if _.text == "0":
            ws["J12"] = "-"
        else:
            if val.is_integer():
                ws["J12"] = int(val)
                ws["J12"].number_format = "#,###"
            else:
                ws["J12"] = round(val, 2)
                ws["J12"].number_format = "#,###.##"

current_path = os.path.dirname(os.path.abspath(__file__))

with open(os.path.join(current_path, "bf_ct.json"), "r", encoding="utf-8") as file:
    bf_ct_json = json.load(file)

tree = xml.etree.ElementTree.parse(arr_immigration)
root = tree.getroot()

arr_country = {}

for _ in root.findall(".//G_IMMIGRATION"):
    fn = _.find("FIRST_NAME").text if _.find("FIRST_NAME") is not None else "-"
    ln = _.find("LAST_NAME").text if _.find("LAST_NAME") is not None else "-"
    nt = _.find("NATIONALITY").text if _.find("NATIONALITY") is not None else "-"
    ct = None
    for nt_imm, cd in bf_ct_json.items():
        if str(nt).strip().lower() == str(nt_imm).strip().lower():
            ct = cd
        elif str(nt).strip().lower() in str(nt_imm).strip().lower():
            ct = cd

    set_name = f"{fn} {ln}"
    arr_country[set_name] = ct

pf = ("PAID", "CTC", "POA", "COA", "POD", "COMP", "MASTER", "COD")

red_color = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

tree = xml.etree.ElementTree.parse(arrival)
root = tree.getroot()

arr_start = 16

for _ in root.findall(".//G_RESERVATION"):
    name = _.find("FULL_NAME_NO_SHR_IND").text if _.find("FULL_NAME_NO_SHR_IND") is not None else "-"
    rm = _.find("DISP_ROOM_NO").text if _.find("DISP_ROOM_NO") is not None else "-"
    ta = _.find("COMPANY_NAME").text if _.find("COMPANY_NAME") is not None else "-"
    vip = _.find("VIP").text if _.find("VIP") is not None else "-"
    memb = None
    for mb in _.findall(".//G_MEM_TYPE_LEVEL"):
        memb = mb.find("MEMBERSHIP_LEVEL").text
    adl = _.find("ADULTS").text if _.find("ADULTS") is not None else "-"
    chd = _.find("CHILDREN").text if _.find("CHILDREN") is not None else "-"
    arr = _.find("ARRIVAL").text if _.find("ARRIVAL") is not None else "-"
    dep = _.find("DEPARTURE").text if _.find("DEPARTURE") is not None else "-"
    eta = _.find("ARRIVAL_TIME").text if _.find("ARRIVAL_TIME") is not None else "-"
    cmt = None
    for cmt_res in _.findall(".//G_COMMENT_RESV_NAME_ID"):
        in_cmt = cmt_res.find("RES_COMMENT").text
        if in_cmt and in_cmt.startswith(pf):
            cmt = in_cmt
        if cmt is None:
            if cmt_res.find("RES_COMMENT_DESCRIPTION").text == "RESERVATION":
                for _ in str(pf).strip().lower():
                        if _ in str(in_cmt).strip().lower():
                            cmt = in_cmt

    set_name = [ _.strip() for _ in name.split(",")]
    title = ["Mr", "Ms", "Mrs", "Miss", "Master", "Dr", "Khun"]
    if len(set_name) == 3:
        ln, fn, tt = set_name[0], set_name[1], set_name[2]
        ws[f"A{arr_start}"] = f"{tt}. {fn} {ln}"
        ws[f"A{arr_start}"].fill = PatternFill(fill_type=None)
        if str(tt).strip().lower() not in str(title).strip().lower():
            ws[f"A{arr_start}"].fill = red_color
    if len(set_name) < 3:
        ln, fn, tt = (set_name + ["", "", ""])[:3]
        ws[f"A{arr_start}"] = f"{tt}. {fn} {ln}"
        ws[f"A{arr_start}"].fill = red_color
    if str(ln).strip() == "" or str(fn).strip() == "" or str(tt).strip() == "":
        ws[f"A{arr_start}"] = f"{tt}. {fn} {ln}"
        ws[f"A{arr_start}"].fill = red_color

    if "Maintenance".strip().lower() in str(name).strip().lower():
        ws[f"A{arr_start}"].fill = red_color

    if f"{fn} {ln}" in arr_country:
        ws[f"C{arr_start}"] = arr_country[f"{fn} {ln}"]
    if ws[f"C{arr_start}"].value is None:
        ws[f"C{arr_start}"] = "-"

    ws[f"D{arr_start}"] = rm

    if "MI Squared".strip().lower() in str(cmt).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nMI Squared"
    if "Booking.com".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nBooking.com"
    if "BG Asia".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nBG Asia"
    if "Siam Tours".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nSiam Tours"
    if "Travelbullz".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nTravelbullz"
    if "BTC".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nBTC"
    if "DERTOUR".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nGo Vacation"
    if "Expedia".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nExpedia"
    if "DNATA".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nDNATA"
    if "Miki".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nMiki Travel"
    if "Rak".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nRak Journeys"
    if "Hotelbeds".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nHotelbeds"
    if "LOTi".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nLOTi"
    if "Ctrip".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nCtrip"
    if "Moon Holidays".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nMoon Holidays"
    if "Martin4T".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nMartin4Travel"
    if "Taecho".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nTaecho"
    if "Trip.com".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nTrip.com"
    if "Vibes Asia".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nVibes Asia"
    if "Pegas".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nPegas"
    if "Trailfinders".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nTrailfinders"
    if "Thailandeal".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nThailandeal"
    if "Thai Center".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nThai Center"
    if "Fly2Thai".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nFly2Thai"
    if "Travel Exclusive".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nTravel Exclusive"
    if "Fun Siam".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nFun Siam"
    if "Elevate Tourism".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nElevate Tourism"
    if "W2M".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nW2M"
    if "Versailles".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nClub Eldorado"
    if "Destinations of".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nDOTW"
    if "DTH".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nDTH"
    if "Y.A.L".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nY.A.L"
    if "Fly East".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nFly East"
    if "Klook Travel".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nKlook Travel"
    if "Meier's".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nMeier's"
    if "MG Group".strip().lower() in str(ta).strip().lower():
        ws[f"D{arr_start}"] = f"{rm}\nMG Group"

    if vip is not None:
        ws[f"F{arr_start}"] = vip
        if memb:
            ws[f"F{arr_start}"] = f"{vip}\n{memb}"

    arr_date = datetime.strptime(arr, "%d/%m/%y")
    dep_date = datetime.strptime(dep, "%d/%m/%y")
    diff_date = dep_date - arr_date

    if vip is None:
        ws[f"F{arr_start}"] = "-"
        if memb in ["SILVER", "GOLD"]:
            if diff_date.days < 7:
                if vip not in ["VIP1", "VIP2", "VIPS"]:
                    ws[f"F{arr_start}"] = f"VIPG\n{memb}"
        if memb in ["PLATINUM"]:
            if vip not in ["VIP1", "VIP2"]:
                ws[f"F{arr_start}"] = f"VIPS\n{memb}"
        if memb in ["TITANIUM", "RED"]:
            if vip not in ["VIP1"]:
                ws[f"F{arr_start}"] = f"VIP2\n{memb}"

    if diff_date.days >= 7:
        if vip not in ["VIP1", "VIP2"]:
            if memb:
                ws[f"F{arr_start}"] = f"VIPS\n{memb}"
            if not memb:
                ws[f"F{arr_start}"] = f"VIPS"

    if memb in ["SILVER", "GOLD"]:
        if diff_date.days < 7:
            if vip not in ["VIP1", "VIP2", "VIPS", "VIPR"]:
                ws[f"F{arr_start}"] = f"VIPG\n{memb}"
                if "repeat".strip().lower() in str(cmt).strip().lower():
                    ws[f"F{arr_start}"] = f"VIPR\n{memb}"
    if memb in ["PLATINUM"]:
        if vip not in ["VIP1", "VIP2"]:
            ws[f"F{arr_start}"] = f"VIPS\n{memb}"
    if memb in ["TITANIUM", "RED"]:
        if vip not in ["VIP1"]:
            ws[f"F{arr_start}"] = f"VIP2\n{memb}"

    if not "MI Squared".strip().lower() in str(cmt).strip().lower():
        if "JADE/RUBY".strip().lower() in str(cmt).strip().lower():
            ws[f"F{arr_start}"] = f"VIPO\nJADE"
        if "DIAMOND".strip().lower() in str(cmt).strip().lower():
            ws[f"F{arr_start}"] = f"VIPO\nDIAMOND"
        if "PLATINUM".strip().lower() in str(cmt).strip().lower():
            ws[f"F{arr_start}"] = f"VIPO\nPLATINUM"
        if "ROYAL".strip().lower() in str(cmt).strip().lower():
            ws[f"F{arr_start}"] = f"VIPO\nROYAL"
    if "MI Squared".strip().lower() in str(cmt).strip().lower():
        if diff_date.days < 7:
            ws[f"F{arr_start}"] = "-"
        if diff_date.days >= 7:
            ws[f"F{arr_start}"] = "VIPS"

    if chd is not None:
        ws[f"H{arr_start}"] = f"{adl}+{chd}"
    if chd is None or chd == "0":
        ws[f"H{arr_start}"] = f"{adl}"

    ws[f"I{arr_start}"] = arr

    ws[f"J{arr_start}"] = dep

    ws[f"K{arr_start}"] = "-"

    if eta is not None:
        ws[f"L{arr_start}"] = eta
        ws[f"L{arr_start}"].value = eta.replace(":", ".")
    if eta is None:
        ws[f"L{arr_start}"] = "15.00"

    ws[f"N{arr_start}"] = "OWN"

    ws[f"P{arr_start}"] = cmt

    arr_start += 1

tree = xml.etree.ElementTree.parse(dep_immigration)
root = tree.getroot()

dep_country = {}

for _ in root.findall(".//G_IMMIGRATION"):
    fn = _.find("FIRST_NAME").text if _.find("FIRST_NAME") is not None else "-"
    ln = _.find("LAST_NAME").text if _.find("LAST_NAME") is not None else "-"
    nt = _.find("NATIONALITY").text if _.find("NATIONALITY") is not None else "-"
    ct = None
    for nt_imm, cd in bf_ct_json.items():
        if str(nt).strip().lower() in str(nt_imm).strip().lower():
            ct = cd

    set_name = f"{fn} {ln}"
    dep_country[set_name] = ct

tree = xml.etree.ElementTree.parse(departure)
root = tree.getroot()

dep_start = arr_start + 2

for _ in root.findall(".//G_ROOM"):
    name = _.find("GUEST_NAME").text if _.find("GUEST_NAME") is not None else "-"
    rm = _.find("ROOM").text if _.find("ROOM") is not None else "-"
    ta = _.find("TRAVEL_AGENT_NAME").text if _.find("TRAVEL_AGENT_NAME") is not None else "-"
    vip = _.find("VIP").text if _.find("VIP") is not None else "-"
    memb = None
    for mb in _.findall(".//LIST_G_MEMBERSHIP/G_MEMBERSHIP"):
        memb = mb.find("MEMBERSHIP_LEVEL").text
    adl = _.find("ADULTS").text if _.find("ADULTS") is not None else "-"
    chd = _.find("CHILDREN").text if _.find("CHILDREN") is not None else "-"
    arr = _.find("CHAR_ARRIVAL").text if _.find("CHAR_ARRIVAL") is not None else "-"
    dep = _.find("CHAR_DEPART").text if _.find("CHAR_DEPART") is not None else "-"
    etd = _.find("DEPARTURE_TIME").text if _.find("DEPARTURE_TIME") is not None else "-"
    cmt = None
    for cmt_res in _.findall("LIST_G_COMMENT_RESV_NAME_ID/G_COMMENT_RESV_NAME_ID"):
        in_cmt = cmt_res.find("RES_COMMENT").text
        if in_cmt and in_cmt.startswith(pf):
            cmt = in_cmt
        if cmt is None:
            if cmt_res.find("RES_COMMENT_DESCRIPTION").text == "Reservation":
                for _ in str(pf).strip().lower():
                    if _ in str(in_cmt).strip().lower():
                        cmt = in_cmt

    set_name = [ _.strip() for _ in name.split(",")]
    title = ["Mr", "Ms", "Mrs", "Miss", "Master", "Dr", "Khun"]
    if len(set_name) == 3:
        ln, fn, tt = set_name[0], set_name[1], set_name[2]
        ws[f"A{dep_start}"] = f"{tt}. {fn} {ln}"
        ws[f"A{dep_start}"].fill = PatternFill(fill_type=None)
        if str(tt).strip().lower() not in str(title).strip().lower():
            ws[f"A{dep_start}"].fill = red_color
    if len(set_name) < 3:
        ln, fn, tt = (set_name + ["", "", ""])[:3]
        ws[f"A{dep_start}"] = f"{tt}. {fn} {ln}"
        ws[f"A{dep_start}"].fill = red_color
    if str(ln).strip() == "" or str(fn).strip() == "" or str(tt).strip() == "":
        ws[f"A{dep_start}"] = f"{tt}. {fn} {ln}"
        ws[f"A{dep_start}"].fill = red_color

    if "Maintenance".strip().lower() in str(name).strip().lower():
        ws[f"A{dep_start}"].fill = red_color

    if f"{fn} {ln}" in dep_country:
        ws[f"C{dep_start}"] = dep_country[f"{fn} {ln}"]
    if ws[f"C{dep_start}"].value is None:
        ws[f"C{dep_start}"] = "-"

    ws[f"D{dep_start}"] = rm

    if "MI Squared".strip().lower() in str(cmt).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nMI Squared"
    if "Booking.com".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nBooking.com"
    if "BG Asia".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nBG Asia"
    if "Siam Tours".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nSiam Tours"
    if "Travelbullz".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nTravelbullz"
    if "BTC".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nBTC"
    if "DERTOUR".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nGo Vacation"
    if "Expedia".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nExpedia"
    if "DNATA".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nDNATA"
    if "Miki".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nMiki Travel"
    if "Rak".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nRak Journeys"
    if "Hotelbeds".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nHotelbeds"
    if "LOTi".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nLOTi"
    if "Ctrip".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nCtrip"
    if "Moon Holidays".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nMoon Holidays"
    if "Martin4T".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nMartin4Travel"
    if "Taecho".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nTaecho"
    if "Trip.com".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nTrip.com"
    if "Vibes Asia".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nVibes Asia"
    if "Pegas".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nPegas"
    if "Trailfinders".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nTrailfinders"
    if "Thailandeal".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nThailandeal"
    if "Thai Center".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nThai Center"
    if "Fly2Thai".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nFly2Thai"
    if "Travel Exclusive".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nTravel Exclusive"
    if "Fun Siam".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nFun Siam"
    if "Elevate Tourism".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nElevate Tourism"
    if "W2M".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nW2M"
    if "Versailles".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nClub Eldorado"
    if "Destinations of".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nDOTW"
    if "DTH".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nDTH"
    if "Y.A.L".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nY.A.L"
    if "Fly East".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nFly East"
    if "Klook Travel".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nKlook Travel"
    if "Meier's".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nMeier's"
    if "MG Group".strip().lower() in str(ta).strip().lower():
        ws[f"D{dep_start}"] = f"{rm}\nMG Group"

    if vip is not None:
        ws[f"F{dep_start}"] = vip
        if memb:
            ws[f"F{dep_start}"] = f"{vip}\n{memb}"

    arr_date = datetime.strptime(arr, "%d/%m/%y")
    dep_date = datetime.strptime(dep, "%d/%m/%y")
    diff_date = dep_date - arr_date

    if vip is None:
        ws[f"F{dep_start}"] = "-"
        if memb in ["SILVER", "GOLD"]:
            if diff_date.days < 7:
                if vip not in ["VIP1", "VIP2", "VIPS"]:
                    ws[f"F{dep_start}"] = f"VIPG\n{memb}"
        if memb in ["PLATINUM"]:
            if vip not in ["VIP1", "VIP2"]:
                ws[f"F{dep_start}"] = f"VIPS\n{memb}"
        if memb in ["TITANIUM", "RED"]:
            if vip not in ["VIP1"]:
                ws[f"F{dep_start}"] = f"VIP2\n{memb}"

    if diff_date.days >= 7:
        if vip not in ["VIP1", "VIP2"]:
            if memb:
                ws[f"F{dep_start}"] = f"VIPS\n{memb}"
            if not memb:
                ws[f"F{dep_start}"] = f"VIPS"

    if memb in ["SILVER", "GOLD"]:
        if diff_date.days < 7:
            if vip not in ["VIP1", "VIP2", "VIPS", "VIPR"]:
                ws[f"F{dep_start}"] = f"VIPG\n{memb}"
                if "repeat".strip().lower() in str(cmt).strip().lower():
                    ws[f"F{dep_start}"] = f"VIPR\n{memb}"
    if memb in ["PLATINUM"]:
        if vip not in ["VIP1", "VIP2"]:
            ws[f"F{dep_start}"] = f"VIPS\n{memb}"
    if memb in ["TITANIUM", "RED"]:
        if vip not in ["VIP1"]:
            ws[f"F{dep_start}"] = f"VIP2\n{memb}"

    if not "MI Squared".strip().lower() in str(cmt).strip().lower():
        if "JADE/RUBY".strip().lower() in str(cmt).strip().lower():
            ws[f"F{dep_start}"] = f"VIPO\nJADE"
        if "DIAMOND".strip().lower() in str(cmt).strip().lower():
            ws[f"F{dep_start}"] = f"VIPO\nDIAMOND"
        if "PLATINUM".strip().lower() in str(cmt).strip().lower():
            ws[f"F{dep_start}"] = f"VIPO\nPLATINUM"
        if "ROYAL".strip().lower() in str(cmt).strip().lower():
            ws[f"F{dep_start}"] = f"VIPO\nROYAL"
    if "MI Squared".strip().lower() in str(cmt).strip().lower():
        if diff_date.days < 7:
            ws[f"F{dep_start}"] = "-"
        if diff_date.days >= 7:
            ws[f"F{dep_start}"] = "VIPS"

    if chd is not None:
        ws[f"H{dep_start}"] = f"{adl}+{chd}"
    if chd is None or chd == "0":
        ws[f"H{dep_start}"] = f"{adl}"

    ws[f"I{dep_start}"] = arr

    ws[f"J{dep_start}"] = dep

    ws[f"K{dep_start}"] = "-"

    if etd is not None:
        ws[f"L{dep_start}"] = etd
        ws[f"L{dep_start}"].value = etd.replace(":", ".")
    if etd is None:
        ws[f"L{dep_start}"] = "12.00"

    ws[f"N{dep_start}"] = "OWN"

    ws[f"P{dep_start}"] = cmt

    dep_start += 1

wb.save(path_td_excel)
wb.close()

subprocess.run(["cmd", "/c", "start", "msedge", f"https://app.reviewpro.com/myPage?fd=2026-01-01&td=2026-08-09&prevFd=2025-01-01&prevTd=2025-08-09&fdManagement=2026-01-01&tdManagement=2026-08-09&lang=en&pid=581134&indexType=GRI&pageId=5d70c1889b6b4944d2ff1bd3"])

if pygetwindow.getWindowsWithTitle("page"):
    pygetwindow.getWindowsWithTitle("page")[0].restore()
    pygetwindow.getWindowsWithTitle("page")[0].maximize()

time.sleep(.5)
script_config.zoom_out(10)
script_config.zoom_in(3)

while True:
    if pyautogui.pixelMatchesColor(1022, 370, (21, 121, 52), tolerance=10):
        break

script_config.zoom_in(7)

pyautogui.press("tab", presses=2, interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.press("tab", interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.press("tab", interval=.01)
pyautogui.press("down", presses=7, interval=.01)
pyautogui.press("enter", interval=.01)
pyautogui.press("tab", presses=7, interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.hotkey("ctrl", "shift", "s", interval=.01)
pyautogui.moveTo(1132, 890)
pyautogui.dragTo(844, 603, button="left")

while True:
    if pyautogui.pixelMatchesColor(1106, 679, (21, 121, 52), tolerance=10):
        break

pyautogui.press("tab", interval=.01)
pyautogui.press("space", interval=.01)

os.startfile(path_td_excel)

if pygetwindow.getWindowsWithTitle("Excel"):
    pygetwindow.getWindowsWithTitle("Excel")[0].activate()
    pygetwindow.getWindowsWithTitle("Excel")[0].maximize()

script_config.stay_excel()

pyautogui.hotkey("ctrl", "g", interval=.01)
pyautogui.write("j1", interval=.01)
pyautogui.press("enter", interval=.01)
pyautogui.hotkey("ctrl", "v", interval=.01)
pyautogui.hotkey("alt", "f11", interval=.01)
pyautogui.press("alt", interval=.01)
pyautogui.press("i", interval=.01)
pyautogui.press("m", interval=.01)
pyautogui.hotkey("ctrl", "a", interval=.01)

gri_mtd = """Sub AutoFitMergeCellPicture()
    Dim pic As Shape
    Dim targetRange As Range
    Dim targetWidth As Double, targetHeight As Double
    
    On Error Resume Next
    For Each pic In Selection.ShapeRange

        If pic.TopLeftCell.MergeCells Then
            Set targetRange = pic.TopLeftCell.MergeArea
        Else
            Set targetRange = pic.TopLeftCell
        End If
        
        targetWidth = targetRange.Width - 4
        targetHeight = targetRange.Height - 4
        
        pic.LockAspectRatio = msoTrue
        
        pic.Width = targetWidth
        If pic.Height > targetHeight Then
            pic.Height = targetHeight
        End If
        
        pic.Width = pic.Width - 100

        pic.Left = targetRange.Left + (targetRange.Width - pic.Width) / 2
        pic.Top = targetRange.Top + (targetRange.Height - pic.Height) - 170
        
    Next pic
End Sub
"""

pyperclip.copy(gri_mtd)
pyautogui.hotkey("ctrl", "v", interval=.01)
pyautogui.press("f5", interval=.01)
pyautogui.hotkey("alt", "f4", interval=.01)
time.sleep(.5)

if pygetwindow.getWindowsWithTitle("page"):
    pygetwindow.getWindowsWithTitle("page")[0].activate()

pyautogui.hotkey("ctrl", "shift", "s", interval=.01)
pyautogui.moveTo(659, 866)
pyautogui.dragTo(334, 584, button="left")
pyautogui.press("tab", interval=.01)
pyautogui.press("space", interval=.01)
time.sleep(.5)

if pygetwindow.getWindowsWithTitle("Excel"):
    pygetwindow.getWindowsWithTitle("Excel")[0].activate()

script_config.stay_excel()

pyautogui.hotkey("ctrl", "g", interval=.01)
pyautogui.write("l1", interval=.01)
pyautogui.press("enter", interval=.01)
pyautogui.hotkey("ctrl", "v", interval=.01)
pyautogui.hotkey("alt", "f11", interval=.01)
pyautogui.hotkey("ctrl", "a", interval=.01)

nps_mtd = """Sub AutoFitMergeCellPicture()
    Dim pic As Shape
    Dim targetRange As Range
    Dim targetWidth As Double, targetHeight As Double
    
    On Error Resume Next
    For Each pic In Selection.ShapeRange

        If pic.TopLeftCell.MergeCells Then
            Set targetRange = pic.TopLeftCell.MergeArea
        Else
            Set targetRange = pic.TopLeftCell
        End If
        
        targetWidth = targetRange.Width - 4
        targetHeight = targetRange.Height - 4
        
        pic.LockAspectRatio = msoTrue
        
        pic.Width = targetWidth
        If pic.Height > targetHeight Then
            pic.Height = targetHeight
        End If
        
        pic.Width = pic.Width - 100

        pic.Left = targetRange.Left + (targetRange.Width - pic.Width) / 2
        pic.Top = targetRange.Top + (targetRange.Height - pic.Height) - 191
        
    Next pic
End Sub
"""

pyperclip.copy(nps_mtd)
pyautogui.hotkey("ctrl", "v", interval=.01)
pyautogui.press("f5", interval=.01)
pyautogui.hotkey("alt", "f4", interval=.01)
time.sleep(.5)

if pygetwindow.getWindowsWithTitle("page"):
    pygetwindow.getWindowsWithTitle("page")[0].activate()

pyautogui.press("space", interval=.01)
pyautogui.press("tab", interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.press("tab", interval=.01)
pyautogui.press("down", presses=14, interval=.01)
pyautogui.press("enter", interval=.01)
pyautogui.press("tab", presses=7, interval=.01)
pyautogui.press("space", interval=.01)
pyautogui.hotkey("ctrl", "shift", "s", interval=.01)
pyautogui.moveTo(1132, 890)
pyautogui.dragTo(844, 603, button="left")

while True:
    if pyautogui.pixelMatchesColor(1106, 679, (21, 121, 52), tolerance=10):
        break

pyautogui.press("tab", interval=.01)
pyautogui.press("space", interval=.01)
time.sleep(.5)

if pygetwindow.getWindowsWithTitle("Excel"):
    pygetwindow.getWindowsWithTitle("Excel")[0].activate()

script_config.stay_excel()

pyautogui.hotkey("ctrl", "g", interval=.01)
pyautogui.write("d1", interval=.01)
pyautogui.press("enter", interval=.01)
pyautogui.hotkey("ctrl", "v", interval=.01)
pyautogui.hotkey("alt", "f11", interval=.01)
pyautogui.hotkey("ctrl", "a", interval=.01)

gri_ytd = """Sub AutoFitMergeCellPicture()
    Dim pic As Shape
    Dim targetRange As Range
    Dim targetWidth As Double, targetHeight As Double
    
    On Error Resume Next
    For Each pic In Selection.ShapeRange

        If pic.TopLeftCell.MergeCells Then
            Set targetRange = pic.TopLeftCell.MergeArea
        Else
            Set targetRange = pic.TopLeftCell
        End If
        
        targetWidth = targetRange.Width - 4
        targetHeight = targetRange.Height - 4
        
        pic.LockAspectRatio = msoTrue
        
        pic.Width = targetWidth
        If pic.Height > targetHeight Then
            pic.Height = targetHeight
        End If
        
        pic.Width = pic.Width - 100

        pic.Left = targetRange.Left + (targetRange.Width - pic.Width) - 50
        pic.Top = targetRange.Top + (targetRange.Height - pic.Height) - 162
        
    Next pic
End Sub
"""

pyperclip.copy(gri_ytd)
pyautogui.hotkey("ctrl", "v", interval=.01)
pyautogui.press("f5", interval=.01)
pyautogui.hotkey("alt", "f4", interval=.01)
time.sleep(.5)

if pygetwindow.getWindowsWithTitle("page"):
    pygetwindow.getWindowsWithTitle("page")[0].activate()

pyautogui.hotkey("ctrl", "shift", "s", interval=.01)
pyautogui.moveTo(659, 866)
pyautogui.dragTo(334, 584, button="left")
pyautogui.press("tab", interval=.01)
pyautogui.press("space", interval=.01)
time.sleep(.5)

if pygetwindow.getWindowsWithTitle("Excel"):
    pygetwindow.getWindowsWithTitle("Excel")[0].activate()

script_config.stay_excel()

pyautogui.hotkey("ctrl", "g", interval=.01)
pyautogui.write("h1", interval=.01)
pyautogui.press("enter", interval=.01)
pyautogui.hotkey("ctrl", "v", interval=.01)
pyautogui.hotkey("alt", "f11", interval=.01)
pyautogui.hotkey("ctrl", "a", interval=.01)

nps_ytd = """Sub AutoFitMergeCellPicture()
    Dim pic As Shape
    Dim targetRange As Range
    Dim targetWidth As Double, targetHeight As Double
    
    On Error Resume Next
    For Each pic In Selection.ShapeRange

        If pic.TopLeftCell.MergeCells Then
            Set targetRange = pic.TopLeftCell.MergeArea
        Else
            Set targetRange = pic.TopLeftCell
        End If
        
        targetWidth = targetRange.Width - 4
        targetHeight = targetRange.Height - 4
        
        pic.LockAspectRatio = msoTrue
        
        pic.Width = targetWidth
        If pic.Height > targetHeight Then
            pic.Height = targetHeight
        End If
        
        pic.Width = pic.Width - 20

        pic.Left = targetRange.Left + (targetRange.Width - pic.Width) / 1.7
        pic.Top = targetRange.Top + (targetRange.Height - pic.Height) - 196
        
    Next pic
End Sub
"""

pyperclip.copy(nps_ytd)
pyautogui.hotkey("ctrl", "v", interval=.01)
pyautogui.press("f5", interval=.01)
pyautogui.hotkey("alt", "f4", interval=.01)
time.sleep(.5)